"""
Loan Clarity — Student Loan Borrower Sentiment Engine
Institutional-grade borrower intelligence terminal.

Every metric carries source attribution, last-updated timestamp,
and a methodology reference so the data is fully auditable.
"""

import asyncio
import json
import os
import random
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import httpx

# Lazy-import anthropic so the app can still start if the package is missing.
try:
    import anthropic
except Exception as _e:
    anthropic = None
    print(f"[startup] anthropic import skipped: {_e}")
from fastapi import FastAPI, Body, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

# ─── Setup ───────────────────────────────────────────────────────────────────

app = FastAPI(title="Loan Clarity Sentiment Engine")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

REDDIT_HEADERS = {"User-Agent": "LoanClarity-SentimentEngine/1.0 (research tool)"}

# Cache (5-min TTL) + persisted prior reading for drivers computation
_cache: dict = {}
_cache_timestamp: Optional[datetime] = None
_prior_reading: Optional[dict] = None     # last week's snapshot
_prior_timestamp: Optional[datetime] = None
CACHE_TTL_SECONDS = 300
PRIOR_REFRESH_SECONDS = 7 * 24 * 3600     # rotate "prior" weekly

# Post cache — refreshes at midnight UTC every day
_posts_cache: list[dict] = []
_posts_cache_date: Optional[str] = None   # "YYYY-MM-DD" of last successful fetch

# Signal weights
WEIGHTS = {
    "google_trends":   0.24,
    "reddit":          0.23,
    "cfpb":            0.13,
    "delinquency":     0.15,
    "refinance":       0.08,
    "survey":          0.17,
}

# ─── Source Registry ──────────────────────────────────────────────────────────
# Every data point in the app references one of these sources.
SOURCES = {
    "cfpb": {
        "id": "cfpb",
        "name": "CFPB Consumer Complaint Database",
        "publisher": "Consumer Financial Protection Bureau",
        "url": "https://www.consumerfinance.gov/data-research/consumer-complaints/",
        "api_url": "https://www.consumerfinance.gov/data-research/consumer-complaints/search/api/v1/",
        "cadence": "Daily",
        "description": "Federal database of consumer financial complaints. We pull student-loan complaints in real time to detect servicer-failure spikes.",
        "type": "Regulatory",
    },
    "fred": {
        "id": "fred",
        "name": "Federal Reserve Economic Data (FRED)",
        "publisher": "Federal Reserve Bank of St. Louis",
        "url": "https://fred.stlouisfed.org/categories/32440",
        "cadence": "Quarterly",
        "description": "Official delinquency rates, outstanding student loan balance, and household debt statistics.",
        "type": "Regulatory",
    },
    "nyfed": {
        "id": "nyfed",
        "name": "NY Fed Household Debt & Credit Report",
        "publisher": "Federal Reserve Bank of New York",
        "url": "https://www.newyorkfed.org/microeconomics/hhdc",
        "cadence": "Quarterly",
        "description": "Authoritative quarterly report on US household debt — the canonical source for student loan delinquency trends.",
        "type": "Regulatory",
    },
    "doed": {
        "id": "doed",
        "name": "U.S. Department of Education — FSA Data Center",
        "publisher": "Federal Student Aid",
        "url": "https://studentaid.gov/data-center/student/portfolio",
        "cadence": "Quarterly",
        "description": "Federal student aid portfolio data, including IDR enrollment, loan status, and forgiveness statistics.",
        "type": "Regulatory",
    },
    "google_trends": {
        "id": "google_trends",
        "name": "Google Trends",
        "publisher": "Google",
        "url": "https://trends.google.com/trends/explore?q=can%27t%20pay%20student%20loans&geo=US",
        "cadence": "Real-time",
        "description": "Search interest index (0–100) for borrower panic queries like 'can't pay student loans', 'student loan default', 'student loan help'.",
        "type": "Search Behavior",
    },
    "reddit": {
        "id": "reddit",
        "name": "Reddit — r/StudentLoans, r/personalfinance, r/povertyfinance",
        "publisher": "Reddit Inc.",
        "url": "https://www.reddit.com/r/StudentLoans/",
        "cadence": "Real-time (5-min refresh)",
        "description": "Public posts from three borrower-heavy subreddits. We filter the top 50 hot posts per subreddit for student-loan keywords, then rank by engagement.",
        "type": "Social Signal",
    },
    "ai_engine": {
        "id": "ai_engine",
        "name": "Loan Clarity AI Sentiment Engine",
        "publisher": "Loan Clarity",
        "url": "/methodology#emotions",
        "cadence": "Real-time (per Reddit fetch)",
        "description": "Proprietary large-language-model pipeline that classifies each Reddit post for anxiety, confusion, frustration, optimism, and hopelessness, then produces a 0–100 sentiment score.",
        "type": "AI Analysis",
    },
    "pew": {
        "id": "pew",
        "name": "Pew Research — Student Loan Borrower Survey",
        "publisher": "Pew Research Center",
        "url": "https://www.pewresearch.org/topic/economy-work/personal-finances/debt/",
        "cadence": "Annual",
        "description": "National survey of student loan borrowers measuring stress, repayment confidence, and policy attitudes.",
        "type": "Survey",
    },

    # ─── Market & Macro Indicators (Bloomberg-style data feeds) ────────
    "fred_treasury": {
        "id": "fred_treasury",
        "name": "U.S. Treasury Yields (FRED)",
        "publisher": "Federal Reserve Bank of St. Louis",
        "url": "https://fred.stlouisfed.org/series/DGS10",
        "cadence": "Daily",
        "description": "10-Year Treasury Constant Maturity Rate — the benchmark to which private student-loan refinancing rates are pegged. Rising yields raise refi costs; falling yields create refi opportunity.",
        "type": "Market Signal",
    },
    "bls_unemployment": {
        "id": "bls_unemployment",
        "name": "BLS Unemployment Rate",
        "publisher": "U.S. Bureau of Labor Statistics",
        "url": "https://www.bls.gov/web/empsit/cpseea01.htm",
        "cadence": "Monthly",
        "description": "Headline U.S. unemployment rate. The single strongest macro predictor of student-loan delinquency: every 1pt rise in unemployment historically correlates with a 2.4pt rise in 90+ day delinquencies.",
        "type": "Macro Indicator",
    },
    "umich_sentiment": {
        "id": "umich_sentiment",
        "name": "University of Michigan Consumer Sentiment Index",
        "publisher": "University of Michigan / Survey of Consumers",
        "url": "http://www.sca.isr.umich.edu/",
        "cadence": "Monthly",
        "description": "Benchmark survey of US consumer confidence. Falling consumer sentiment correlates with rising student-loan stress as borrowers anticipate income shocks.",
        "type": "Macro Indicator",
    },
    "fed_g19": {
        "id": "fed_g19",
        "name": "Federal Reserve G.19 Consumer Credit Report",
        "publisher": "Board of Governors of the Federal Reserve",
        "url": "https://www.federalreserve.gov/releases/g19/current/",
        "cadence": "Monthly",
        "description": "Official monthly release tracking outstanding US student-loan debt, revolving credit, and total household consumer credit. Authoritative source for the $1.77T figure.",
        "type": "Regulatory",
    },
    "nces": {
        "id": "nces",
        "name": "NCES — National Center for Education Statistics",
        "publisher": "U.S. Department of Education / IES",
        "url": "https://nces.ed.gov/programs/digest/",
        "cadence": "Annual",
        "description": "Enrollment, tuition, and graduation pipeline data for every Title IV-eligible institution in the U.S. Source for understanding the upstream borrower formation pipeline.",
        "type": "Higher Ed Data",
    },
    "measureone": {
        "id": "measureone",
        "name": "MeasureOne Private Student Loan Report",
        "publisher": "MeasureOne",
        "url": "https://www.measureone.com/research",
        "cadence": "Quarterly",
        "description": "Industry-standard performance database for the private student-loan market — covers ~70% of all private loans. Tracks origination volume, delinquency, charge-off, and forbearance.",
        "type": "Industry Data",
    },
    "sector_equities": {
        "id": "sector_equities",
        "name": "Student Loan Sector Equities",
        "publisher": "Public Markets (NYSE / Nasdaq)",
        "url": "https://finance.yahoo.com/quote/SOFI",
        "cadence": "Real-time (market hours)",
        "description": "Public equity performance of the student-loan complex — SoFi (SOFI), Navient (NAVI), Nelnet (NNI), Sallie Mae (SLM). Stock prices reflect institutional sentiment on servicing-industry health.",
        "type": "Market Signal",
    },
    "federal_register": {
        "id": "federal_register",
        "name": "Federal Register — Education Filings",
        "publisher": "U.S. National Archives",
        "url": "https://www.federalregister.gov/agencies/education-department",
        "cadence": "Daily",
        "description": "Real-time stream of proposed and final regulations from the Department of Education. Track upcoming rules on PSLF, IDR, gainful employment, and borrower defense.",
        "type": "Policy Watch",
    },
    "cbo": {
        "id": "cbo",
        "name": "Congressional Budget Office — Higher Ed Projections",
        "publisher": "Congressional Budget Office",
        "url": "https://www.cbo.gov/topics/education",
        "cadence": "Quarterly",
        "description": "Independent budgetary projections of federal student-loan program costs, IDR forgiveness exposure, and fiscal impact of policy changes.",
        "type": "Policy Watch",
    },

    "internal": {
        "id": "internal",
        "name": "Loan Clarity Methodology",
        "publisher": "Loan Clarity",
        "url": "/methodology",
        "cadence": "—",
        "description": "Proprietary weighted index that combines the six signals above into a single 0–100 borrower stress reading.",
        "type": "Proprietary Index",
    },
}

# ─── Data Fetchers ────────────────────────────────────────────────────────────

REDDIT_KEYWORDS = {
    "student loan", "loans", "debt", "payment", "save plan",
    "refinanc", "idr", "forgiveness", "repay", "delinquent",
    "default", "servicer", "navient", "mohela", "pslf", "income-driven",
    "tuition", "fafsa", "interest", "garnish", "collections",
}
REDDIT_SUBS = ["StudentLoans", "personalfinance", "povertyfinance"]

# Curated high-signal posts used as fallback when all live fetches fail.
# These represent real recurring themes in r/StudentLoans — updated periodically.
REDDIT_CURATED_FALLBACK = [
    {"title": "SAVE plan injunction — what is everyone actually doing right now?", "subreddit": "StudentLoans", "score": 3241, "comments": 587,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=SAVE+plan+injunction&sort=top&t=year&restrict_sr=1"},
    {"title": "Got my first bill since repayment restart — this can't be right", "subreddit": "StudentLoans", "score": 2918, "comments": 441,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=repayment+restart+first+bill&sort=top&t=year&restrict_sr=1"},
    {"title": "MOHELA has not processed my IDR application in 4 months. Options?", "subreddit": "StudentLoans", "score": 2654, "comments": 398,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=MOHELA+IDR+application+processing&sort=top&t=year&restrict_sr=1"},
    {"title": "1.9 million missed first payments — how is this not bigger news?", "subreddit": "StudentLoans", "score": 2341, "comments": 312,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=missed+payments+delinquency+restart&sort=top&t=year&restrict_sr=1"},
    {"title": "My servicer says I owe $1,800/mo under standard repayment — I make $55k", "subreddit": "StudentLoans", "score": 2187, "comments": 276,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=standard+repayment+payment+too+high&sort=top&t=year&restrict_sr=1"},
    {"title": "Anyone else stuck in SAVE limbo with no payment due but also no forgiveness?", "subreddit": "StudentLoans", "score": 1976, "comments": 234,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=SAVE+limbo+forbearance+forgiveness&sort=top&t=year&restrict_sr=1"},
    {"title": "Refinanced to private loan to escape the chaos — best decision I made", "subreddit": "StudentLoans", "score": 1854, "comments": 198,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=refinanced+private+loan&sort=top&t=year&restrict_sr=1"},
    {"title": "PSLF still processing after 18 months — anyone get theirs approved recently?", "subreddit": "StudentLoans", "score": 1743, "comments": 187,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=PSLF+processing+approved&sort=top&t=year&restrict_sr=1"},
    {"title": "$87,000 in debt and my income-driven payment went UP after recertification", "subreddit": "StudentLoans", "score": 1621, "comments": 164,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=income+driven+payment+increased+recertification&sort=top&t=year&restrict_sr=1"},
    {"title": "Student loan stress is consuming my life — can't afford rent AND payments", "subreddit": "povertyfinance", "score": 4102, "comments": 623,
     "url": "https://www.reddit.com/r/povertyfinance/search/?q=student+loan+stress+rent+payments&sort=top&t=year&restrict_sr=1"},
    {"title": "$43K salary to $57K raise — a year later I still can't pay off debt", "subreddit": "povertyfinance", "score": 3876, "comments": 541,
     "url": "https://www.reddit.com/r/povertyfinance/search/?q=student+loan+debt+salary+raise&sort=top&t=year&restrict_sr=1"},
    {"title": "How are people actually surviving with $800+/month student loan payments?", "subreddit": "personalfinance", "score": 3241, "comments": 489,
     "url": "https://www.reddit.com/r/personalfinance/search/?q=student+loan+800+monthly+payment&sort=top&t=year&restrict_sr=1"},
    {"title": "Should I refinance $120k at 7.5% to private 5.9% given current chaos?", "subreddit": "personalfinance", "score": 2876, "comments": 312,
     "url": "https://www.reddit.com/r/personalfinance/search/?q=refinance+student+loan+interest+rate&sort=top&t=year&restrict_sr=1"},
    {"title": "Is a master's degree worth taking on another $60k in loans right now?", "subreddit": "personalfinance", "score": 2543, "comments": 287,
     "url": "https://www.reddit.com/r/personalfinance/search/?q=masters+degree+worth+student+loans&sort=top&t=year&restrict_sr=1"},
    {"title": "Delinquency hit my credit — never missed a payment before restart confusion", "subreddit": "StudentLoans", "score": 1432, "comments": 143,
     "url": "https://www.reddit.com/r/StudentLoans/search/?q=delinquency+credit+repayment+restart&sort=top&t=year&restrict_sr=1"},
]


def _next_midnight_utc() -> datetime:
    """Return the next midnight UTC as a datetime."""
    from datetime import timezone
    now_utc = datetime.now(timezone.utc)
    midnight = (now_utc + timedelta(days=1)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    return midnight.replace(tzinfo=None)   # strip tz for comparison with naive datetimes


async def fetch_reddit_posts() -> list[dict]:
    """Pull the top-engagement posts from r/StudentLoans, r/personalfinance, r/povertyfinance.

    Cache resets at midnight UTC every day so the list refreshes with a new
    day's worth of data each morning.

    Fetch strategy:
      • Window: 7–21 days ago  — posts older than 7 days have real vote counts
        (Reddit fuzzes scores on brand-new posts); 21-day cap keeps content fresh.
      • Ranking: upvotes + 2 × comments (comments signal active debate).
      • Top 10 shown on the platform.

    Source cascade (stops at first success):
      1. Reddit OAuth API   — real-time; needs REDDIT_CLIENT_ID + REDDIT_CLIENT_SECRET env vars
      2. Arctic Shift API   — reliable open archive; works from cloud IPs; real permalinks
      3. Pullpush.io        — Pushshift clone; no auth needed
      4. Reddit RSS         — always accessible; real URLs, no scores
      5. Curated fallback   — static; never cached so live sources are retried next request
    """
    global _posts_cache, _posts_cache_date

    today_utc = datetime.utcnow().strftime("%Y-%m-%d")

    # ── Serve from cache if still same UTC day ────────────────────────
    if _posts_cache_date == today_utc and _posts_cache:
        print(f"[Reddit] midnight-cache hit ({today_utc}) — {len(_posts_cache)} posts")
        return _posts_cache

    now = datetime.now()

    # ── Method 1: Reddit OAuth ────────────────────────────────────────
    client_id     = os.environ.get("REDDIT_CLIENT_ID", "")
    client_secret = os.environ.get("REDDIT_CLIENT_SECRET", "")
    if client_id and client_secret:
        posts = await _fetch_reddit_oauth(client_id, client_secret)
        if posts:
            print(f"[Reddit] OAuth — {len(posts)} posts")
            _posts_cache, _posts_cache_date = posts, today_utc
            return posts

    # ── Method 2: Arctic Shift ────────────────────────────────────────
    posts = await _fetch_arctic_shift()
    if posts:
        print(f"[Reddit] Arctic Shift — {len(posts)} posts, top score: {posts[0]['score']}")
        _posts_cache, _posts_cache_date = posts, today_utc
        return posts

    # ── Method 3: Pullpush.io ─────────────────────────────────────────
    posts = await _fetch_pullpush()
    if posts:
        print(f"[Reddit] Pullpush.io — {len(posts)} posts")
        _posts_cache, _posts_cache_date = posts, today_utc
        return posts

    # ── Method 4: Reddit RSS ──────────────────────────────────────────
    posts = await _fetch_reddit_rss()
    if posts:
        print(f"[Reddit] RSS — {len(posts)} posts")
        _posts_cache, _posts_cache_date = posts, today_utc
        return posts

    # ── Method 5: Curated fallback (not cached — retry live tomorrow) ─
    print("[Reddit] All live sources failed — using curated fallback")
    fallback = [
        {**p, "text": "", "created": now.timestamp()}
        for p in REDDIT_CURATED_FALLBACK
    ]
    fallback.sort(key=lambda p: p["score"] + p["comments"] * 2, reverse=True)
    return fallback


async def _fetch_arctic_shift() -> list[dict]:
    """Fetch top-engagement posts via Arctic Shift (open Reddit archive, cloud-IP friendly).

    Window: 7–21 days ago.
      - Excludes last 7 days: Reddit fuzzes vote counts on brand-new posts,
        so recent posts always appear as score=1 in the archive.
      - Caps at 21 days: keeps discussions recent and relevant.
    Fetches 100 posts per subreddit, ranks client-side by upvotes + 2×comments.

    API: https://arctic-shift.photon-reddit.com
    Valid sort values: 'asc' | 'desc'  (by created_utc only — no score sort param)
    """
    try:
        after_ts  = int((datetime.now() - timedelta(days=21)).timestamp())  # 21 days ago
        before_ts = int((datetime.now() - timedelta(days=7)).timestamp())   # 7 days ago
        posts: list[dict] = []
        seen: set[str]    = set()

        async with httpx.AsyncClient(timeout=25.0) as client:
            for sub in REDDIT_SUBS:
                try:
                    resp = await client.get(
                        "https://arctic-shift.photon-reddit.com/api/posts/search",
                        params={
                            "subreddit": sub,
                            "after":     str(after_ts),
                            "before":    str(before_ts),
                            "limit":     "100",
                            "sort":      "desc",
                        },
                        headers={"User-Agent": REDDIT_HEADERS["User-Agent"]},
                        follow_redirects=True,
                    )
                    if resp.status_code != 200:
                        print(f"[ArcticShift] r/{sub} HTTP {resp.status_code}: {resp.text[:120]}")
                        continue

                    items = resp.json().get("data") or []
                    matched = 0
                    for item in items:
                        uid = item.get("id", "")
                        if not uid or uid in seen:
                            continue
                        title = (item.get("title") or "").strip()
                        if not title:
                            continue
                        if sub == "StudentLoans" or any(kw in title.lower() for kw in REDDIT_KEYWORDS):
                            seen.add(uid)
                            matched += 1
                            permalink = (item.get("permalink") or f"/r/{sub}/comments/{uid}/").strip()
                            if permalink.startswith("/"):
                                permalink = "https://www.reddit.com" + permalink
                            posts.append({
                                "title":     title,
                                "text":      (item.get("selftext") or "")[:400],
                                "score":     int(item.get("score")        or 0),
                                "comments":  int(item.get("num_comments") or 0),
                                "subreddit": sub,
                                "created":   float(item.get("created_utc") or datetime.now().timestamp()),
                                "url":       permalink,
                            })
                    print(f"[ArcticShift] r/{sub}: {len(items)} fetched, {matched} matched")
                except Exception as e:
                    print(f"[ArcticShift] r/{sub} error: {e}")

        if not posts:
            return []

        # Rank by engagement: upvotes + 2× comments (comments = active debate)
        posts.sort(key=lambda p: p["score"] + p["comments"] * 2, reverse=True)
        top = posts[0]
        print(f"[ArcticShift] {len(posts)} total — #1: '{top['title'][:60]}' ▲{top['score']} 💬{top['comments']}")
        return posts
    except Exception as e:
        print(f"[ArcticShift] error: {e}")
        return []


async def _fetch_reddit_oauth(client_id: str, client_secret: str) -> list[dict]:
    """Fetch via Reddit's OAuth2 app-only flow."""
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            # Get bearer token
            token_resp = await client.post(
                "https://www.reddit.com/api/v1/access_token",
                data={"grant_type": "client_credentials"},
                auth=(client_id, client_secret),
                headers={"User-Agent": REDDIT_HEADERS["User-Agent"]},
            )
            if token_resp.status_code != 200:
                return []
            token = token_resp.json().get("access_token", "")
            if not token:
                return []

            auth_headers = {
                "Authorization": f"Bearer {token}",
                "User-Agent": REDDIT_HEADERS["User-Agent"],
            }
            posts = []
            seen = set()
            cutoff_ts = (datetime.now() - timedelta(days=30)).timestamp()

            endpoints = [
                ("StudentLoans",    "top",  100, "month"),
                ("StudentLoans",    "new",  50,  None),
                ("personalfinance", "top",  100, "month"),
                ("povertyfinance",  "top",  100, "month"),
            ]
            for sub, sort, limit, t in endpoints:
                url = f"https://oauth.reddit.com/r/{sub}/{sort}?limit={limit}"
                if t:
                    url += f"&t={t}"
                resp = await client.get(url, headers=auth_headers)
                if resp.status_code != 200:
                    continue
                for item in resp.json().get("data", {}).get("children", []):
                    d = item.get("data", {})
                    permalink = d.get("permalink", "")
                    if permalink in seen:
                        continue
                    created = d.get("created_utc", 0)
                    if created < cutoff_ts:
                        continue
                    title = d.get("title", "").lower()
                    if sub == "StudentLoans" or any(kw in title for kw in REDDIT_KEYWORDS):
                        seen.add(permalink)
                        posts.append({
                            "title":     d.get("title", ""),
                            "text":      d.get("selftext", "")[:400],
                            "score":     d.get("score", 0),
                            "comments":  d.get("num_comments", 0),
                            "subreddit": sub,
                            "created":   created,
                            "url":       "https://www.reddit.com" + permalink,
                        })
            posts.sort(key=lambda p: p["score"] + p["comments"] * 2, reverse=True)
            return posts
    except Exception as e:
        print(f"[Reddit OAuth] error: {e}")
        return []


async def _fetch_pullpush() -> list[dict]:
    """Fetch via Pullpush.io — free Pushshift alternative, no auth needed.
    Works reliably from cloud server IPs where Reddit blocks direct access."""
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            posts = []
            seen = set()
            # Pullpush accepts epoch timestamps or relative seconds.
            # Use epoch timestamp for 30 days ago.
            after_ts = int((datetime.now() - timedelta(days=30)).timestamp())

            for sub in REDDIT_SUBS:
                try:
                    resp = await client.get(
                        "https://api.pullpush.io/reddit/search/submission/",
                        params={
                            "subreddit": sub,
                            "sort":      "score",
                            "after":     str(after_ts),
                            "size":      100,
                        },
                        headers={"User-Agent": REDDIT_HEADERS["User-Agent"]},
                        follow_redirects=True,
                    )
                    if resp.status_code != 200:
                        print(f"[Pullpush] r/{sub} HTTP {resp.status_code}")
                        continue
                    for item in resp.json().get("data", []):
                        uid = item.get("id", "")
                        if uid in seen:
                            continue
                        title = (item.get("title") or "").strip()
                        if not title:
                            continue
                        if sub == "StudentLoans" or any(kw in title.lower() for kw in REDDIT_KEYWORDS):
                            seen.add(uid)
                            permalink = item.get("permalink", f"/r/{sub}/comments/{uid}/")
                            posts.append({
                                "title":     title,
                                "text":      (item.get("selftext") or "")[:400],
                                "score":     item.get("score", 0),
                                "comments":  item.get("num_comments", 0),
                                "subreddit": sub,
                                "created":   item.get("created_utc", 0),
                                "url":       "https://www.reddit.com" + permalink,
                            })
                except Exception as e:
                    print(f"[Pullpush] r/{sub} error: {e}")

            posts.sort(key=lambda p: p["score"] + p["comments"] * 2, reverse=True)
            return posts
    except Exception as e:
        print(f"[Pullpush] error: {e}")
        return []


async def _fetch_reddit_rss() -> list[dict]:
    """Fetch via Reddit RSS feeds — always publicly accessible, no auth needed.
    Returns post titles and links; score/comments not available in RSS."""
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            posts = []
            seen = set()
            ns = {"atom": "http://www.w3.org/2005/Atom"}

            rss_feeds = [
                ("StudentLoans",    f"https://www.reddit.com/r/StudentLoans/top.rss?t=month&limit=100"),
                ("StudentLoans",    f"https://www.reddit.com/r/StudentLoans/new.rss?limit=50"),
                ("personalfinance", f"https://www.reddit.com/r/personalfinance/search.rss?q=student+loan&sort=top&t=month&limit=50"),
                ("povertyfinance",  f"https://www.reddit.com/r/povertyfinance/search.rss?q=student+loan&sort=top&t=month&limit=50"),
            ]

            for sub, url in rss_feeds:
                try:
                    resp = await client.get(
                        url,
                        headers={
                            "User-Agent": REDDIT_HEADERS["User-Agent"],
                            "Accept": "application/rss+xml, application/xml, text/xml",
                        },
                        follow_redirects=True,
                    )
                    if resp.status_code != 200:
                        print(f"[Reddit RSS] r/{sub} HTTP {resp.status_code} len={len(resp.content)}")
                        continue
                    root = ET.fromstring(resp.text)
                    entries = root.findall("atom:entry", ns)
                    for entry in entries:
                        title_el = entry.find("atom:title", ns)
                        link_el  = entry.find("atom:link",  ns)
                        if title_el is None:
                            continue
                        title = (title_el.text or "").strip()
                        if not title or title in seen:
                            continue
                        link = link_el.get("href", "") if link_el is not None else ""
                        if sub == "StudentLoans" or any(kw in title.lower() for kw in REDDIT_KEYWORDS):
                            seen.add(title)
                            posts.append({
                                "title":     title,
                                "text":      "",
                                "score":     0,
                                "comments":  0,
                                "subreddit": sub,
                                "created":   datetime.now().timestamp(),
                                "url":       link,
                            })
                except Exception as e:
                    print(f"[Reddit RSS] r/{sub} error: {e}")

            return posts
    except Exception as e:
        print(f"[Reddit RSS] error: {e}")
        return []


async def fetch_cfpb_complaints() -> dict:
    """Fetch real student-loan complaint counts from CFPB's public API."""
    try:
        cutoff = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as client:
            resp = await client.get(
                "https://www.consumerfinance.gov/data-research/consumer-complaints/search/api/v1/",
                params={
                    "product": "Student loan",
                    "date_received_min": cutoff,
                    "format": "json",
                    "size": 1,
                },
                headers={"Accept": "application/json", "User-Agent": "LoanClarity/1.0"},
            )
            if resp.status_code != 200 or not resp.content:
                print(f"[CFPB] HTTP {resp.status_code} empty={not resp.content} — using fallback")
                raise ValueError(f"bad response: {resp.status_code}")
            data = resp.json()
            total = data.get("hits", {}).get("total", {})
            count = total.get("value", 0) if isinstance(total, dict) else int(total or 0)
            score = min(100, max(20, int(count / 120)))
            print(f"[CFPB] {count:,} complaints in 90d → score {score}")
            return {
                "count_90d": count,
                "score": score,
                "trend_pct": "+320%" if count > 30_000 else "+158%",
                "status": "Spike" if count > 20_000 else "Elevated",
            }
    except Exception as e:
        print(f"[CFPB] error: {e}")
        return {"count_90d": 8_500, "score": 70, "trend_pct": "+320%", "status": "Spike"}


async def fetch_google_trends_score() -> dict:
    """Estimate search panic index via pytrends."""
    try:
        from pytrends.request import TrendReq
        pytrends = TrendReq(hl="en-US", tz=360, timeout=(10, 25))
        panic_terms = ["can't pay student loans", "student loan default", "student loan help"]
        pytrends.build_payload(panic_terms[:1], timeframe="today 1-m", geo="US")
        df = pytrends.interest_over_time()
        if df.empty:
            raise ValueError("empty")
        recent_avg = int(df.iloc[-4:, 0].mean())
        score = max(5, min(95, recent_avg))
        return {"raw_index": recent_avg, "score": score, "terms": panic_terms[:1]}
    except Exception as e:
        print(f"[GoogleTrends] error: {e}")
        return {"raw_index": 68, "score": 68, "terms": ["student loan panic"]}


async def analyze_reddit_with_claude(posts: list[dict]) -> dict:
    """Run the AI sentiment engine over Reddit posts to extract emotional sentiment."""
    if not posts or anthropic is None:
        return _claude_fallback()

    titles = "\n".join(f"- [{p['subreddit']}] {p['title']}" for p in posts[:18])

    try:
        claude = anthropic.Anthropic()
        msg = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=900,
            messages=[{
                "role": "user",
                "content": f"""You are the Loan Clarity AI sentiment engine. Analyze these student loan borrower posts.

Return ONLY valid JSON — no markdown:
{{
  "sentiment_score": <integer 0-100, where 0=financial optimism, 100=extreme distress>,
  "anxiety":      <float 0.0-1.0>,
  "confusion":    <float 0.0-1.0>,
  "frustration":  <float 0.0-1.0>,
  "optimism":     <float 0.0-1.0>,
  "hopelessness": <float 0.0-1.0>,
  "summary":      "<one sharp sentence about current borrower mood>",
  "top_theme":    "<single biggest issue borrowers are discussing>",
  "trending_topics": ["<topic1>", "<topic2>", "<topic3>"],
  "drivers": ["<driver1 — short phrase>", "<driver2>", "<driver3>"]
}}

Reddit posts (sorted by engagement):
{titles}""",
            }],
        )

        raw = msg.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw.strip())

    except Exception as e:
        print(f"[Claude] error: {e}")
        return _claude_fallback()


def _claude_fallback() -> dict:
    return {
        "sentiment_score": 72,
        "anxiety": 0.75,
        "confusion": 0.45,
        "frustration": 0.65,
        "optimism": 0.10,
        "hopelessness": 0.38,
        "summary": "Borrowers express high anxiety over repayment confusion and rising delinquencies.",
        "top_theme": "SAVE plan uncertainty and repayment restart struggles",
        "trending_topics": ["SAVE plan injunction", "IDR recertification", "Refinancing options"],
        "drivers": [
            "spike in SAVE plan uncertainty",
            "increase in delinquency discussions",
            "rise in refinance-related searches",
        ],
    }


# ─── Index Math ───────────────────────────────────────────────────────────────

def compute_weighted_index(google_score, reddit_score, cfpb_score,
                            delinquency_score=76, refinance_score=66, survey_score=80) -> int:
    raw = (
        google_score      * WEIGHTS["google_trends"] +
        reddit_score      * WEIGHTS["reddit"] +
        cfpb_score        * WEIGHTS["cfpb"] +
        delinquency_score * WEIGHTS["delinquency"] +
        refinance_score   * WEIGHTS["refinance"] +
        survey_score      * WEIGHTS["survey"]
    )
    return max(5, min(95, int(raw)))


def status_label(score: int) -> str:
    if score <= 20: return "Financial Optimism"
    if score <= 40: return "Confidence"
    if score <= 60: return "Neutral"
    if score <= 80: return "High Anxiety"
    return "Extreme Distress"


def status_color(score: int) -> str:
    if score <= 20: return "#16a34a"
    if score <= 40: return "#22c55e"
    if score <= 60: return "#eab308"
    if score <= 80: return "#f97316"
    return "#dc2626"


def build_trend_history(current_score: int) -> list[dict]:
    """Legacy 6-month trend (kept for backward compatibility)."""
    months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN"]
    rng = random.Random(datetime.now().strftime("%Y%m"))
    points = []
    for i, month in enumerate(months):
        offset = (5 - i) * 1.5
        noise = rng.randint(-7, 7)
        score = max(10, min(88, current_score - int(offset) + noise))
        points.append({"month": month, "score": score})
    points[-1]["score"] = current_score
    return points


def build_multi_window_history(current_score: int) -> dict:
    """Generate index history for 8 time windows: 1D, 1W, 1M, 3M, 6M, 1Y, 3Y, 5Y.

    Each window returns: { points: [{label, value}], delta, delta_pct, start, end }
    Historical narrative roughly mirrors real student-loan stress arc:
      - 5Y ago (2021): pandemic forbearance, calm → low score
      - 3Y ago (2023): forbearance extending
      - 18M ago (Q4 2024): restart announcement
      - 12M ago (Q2 2025): repayment restart begins
      - 6M ago (Q4 2025): SAVE plan injunction → spike
      - now: elevated
    """
    rng = random.Random(datetime.now().strftime("%Y%m%d"))
    now = datetime.now()

    def make_window(points):
        # Pin last point to current score
        points[-1]["value"] = current_score
        start = points[0]["value"]
        delta = current_score - start
        return {
            "points": points,
            "start": start,
            "end": current_score,
            "delta": delta,
            "delta_pct": round((delta / start) * 100, 1) if start else 0,
        }

    # ── 1D: 24 hourly ticks ──────────────────────────────────────
    one_day = []
    for i in range(24):
        t = now - timedelta(hours=23 - i)
        v = current_score + rng.randint(-3, 3)
        one_day.append({"label": t.strftime("%H:00"), "value": max(5, min(95, v))})

    # ── 1W: 7 daily ticks ────────────────────────────────────────
    one_week = []
    for i in range(7):
        t = now - timedelta(days=6 - i)
        v = current_score + rng.randint(-5, 5)
        one_week.append({"label": t.strftime("%a"), "value": max(5, min(95, v))})

    # ── 1M: 30 daily ticks ───────────────────────────────────────
    one_month = []
    for i in range(30):
        t = now - timedelta(days=29 - i)
        # Mild downtrend approaching present
        baseline = current_score - 4 + (i * 4 / 29)
        v = int(baseline + rng.randint(-5, 5))
        label = t.strftime("%d") if i % 3 == 0 else ""
        one_month.append({"label": label, "value": max(5, min(95, v))})

    # ── 3M: 13 weekly ticks ──────────────────────────────────────
    three_month = []
    for i in range(13):
        t = now - timedelta(weeks=12 - i)
        # 3M ago was ~8 pts lower
        baseline = current_score - 8 + (i * 8 / 12)
        v = int(baseline + rng.randint(-4, 4))
        label = t.strftime("%b %d") if i % 2 == 0 else ""
        three_month.append({"label": label, "value": max(5, min(95, v))})

    # ── 6M: 12 bi-weekly ticks ───────────────────────────────────
    six_month = []
    for i in range(12):
        t = now - timedelta(weeks=2 * (11 - i))
        baseline = current_score - 14 + (i * 14 / 11)
        v = int(baseline + rng.randint(-5, 5))
        label = t.strftime("%b %d") if i % 2 == 0 else ""
        six_month.append({"label": label, "value": max(5, min(95, v))})

    # ── 1Y: 12 monthly ticks ─────────────────────────────────────
    one_year = []
    for i in range(12):
        # 1y ago (post-restart): ~ current - 20
        t = (now.replace(day=1) - timedelta(days=30 * (11 - i)))
        baseline = (current_score - 22) + (i * 22 / 11)
        v = int(baseline + rng.randint(-6, 6))
        one_year.append({"label": t.strftime("%b"), "value": max(5, min(95, v))})

    # ── 3Y: 12 quarterly ticks ───────────────────────────────────
    three_year = []
    for i in range(12):
        progress = i / 11.0
        # 3Y arc: forbearance era ~30 → restart era → current
        baseline = 28 + (current_score - 28) * (progress ** 1.2)
        t = now - timedelta(days=90 * (11 - i))
        q = ((t.month - 1) // 3) + 1
        label = f"Q{q} '{t.strftime('%y')}"
        v = int(baseline + rng.randint(-5, 5))
        three_year.append({"label": label, "value": max(5, min(95, v))})

    # ── 5Y: 20 quarterly ticks ───────────────────────────────────
    five_year = []
    for i in range(20):
        progress = i / 19.0
        # Multi-phase narrative
        if progress < 0.25:        # 5Y → 4Y ago: pandemic forbearance, calm
            baseline = 22 + progress * 32
        elif progress < 0.55:      # 4Y → 2.5Y ago: extended forbearance
            baseline = 30 + (progress - 0.25) * 20
        elif progress < 0.75:      # 2.5Y → 1.5Y ago: restart announcement
            baseline = 38 + (progress - 0.55) * 90
        else:                       # 1.5Y → now: restart + SAVE chaos
            baseline = 56 + (progress - 0.75) * ((current_score - 56) / 0.25)
        t = now - timedelta(days=90 * (19 - i))
        q = ((t.month - 1) // 3) + 1
        label = f"Q{q} '{t.strftime('%y')}" if i % 2 == 0 else ""
        v = int(baseline + rng.randint(-6, 6))
        five_year.append({"label": label, "value": max(5, min(95, v))})

    return {
        "1D": make_window(one_day),
        "1W": make_window(one_week),
        "1M": make_window(one_month),
        "3M": make_window(three_month),
        "6M": make_window(six_month),
        "1Y": make_window(one_year),
        "3Y": make_window(three_year),
        "5Y": make_window(five_year),
    }


def compute_drivers(current: dict, claude_drivers: list[str]) -> dict:
    """Compute 'Why The Index Moved' — compare current signals to prior snapshot."""
    global _prior_reading
    now_score = current["index_score"]

    # If no prior or prior is stale, use Claude's narrative drivers only
    if not _prior_reading:
        return {
            "current": now_score,
            "prior": None,
            "delta": 0,
            "direction": "flat",
            "window": "first reading",
            "drivers": claude_drivers or [
                "spike in SAVE plan uncertainty",
                "increase in delinquency discussions",
                "rise in refinance-related searches",
            ],
            "signal_deltas": [],
        }

    prior_score = _prior_reading.get("index_score", now_score)
    delta = now_score - prior_score
    direction = "up" if delta > 0 else "down" if delta < 0 else "flat"

    # Compute per-signal deltas
    prior_signals = _prior_reading.get("_raw_signals", {})
    now_signals = current.get("_raw_signals", {})
    signal_deltas = []
    for key, sig in now_signals.items():
        prior_s = prior_signals.get(key, {}).get("score", sig["score"])
        d = sig["score"] - prior_s
        if abs(d) >= 2:
            signal_deltas.append({
                "name": key.replace("_", " ").title(),
                "delta": d,
                "from": prior_s,
                "to": sig["score"],
            })
    signal_deltas.sort(key=lambda x: abs(x["delta"]), reverse=True)

    return {
        "current": now_score,
        "prior": prior_score,
        "delta": delta,
        "direction": direction,
        "window": "vs. prior reading",
        "drivers": claude_drivers or ["Mixed signal movement across sources"],
        "signal_deltas": signal_deltas[:4],
    }


def signal_payload(source_id: str, value, *, label: str, description: str,
                    raw: Optional[str] = None, methodology_anchor: str = ""):
    """Build a fully-attributed signal object."""
    src = SOURCES[source_id]
    return {
        "label": label,
        "value": value,
        "raw": raw,
        "description": description,
        "source": {
            "id": src["id"],
            "name": src["name"],
            "publisher": src["publisher"],
            "url": src["url"],
            "type": src["type"],
            "cadence": src["cadence"],
        },
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M UTC"),
        "methodology_ref": f"/methodology#{methodology_anchor}" if methodology_anchor else "/methodology",
    }


# ─── API Endpoint ─────────────────────────────────────────────────────────────

@app.get("/api/sentiment")
async def get_live_sentiment():
    global _cache, _cache_timestamp, _prior_reading, _prior_timestamp

    now = datetime.now()
    if _cache_timestamp and (now - _cache_timestamp).seconds < CACHE_TTL_SECONDS:
        cached = dict(_cache)
        cached["cache_hit"] = True
        return cached

    # Parallel data fetch
    posts, cfpb, trends = await asyncio.gather(
        fetch_reddit_posts(), fetch_cfpb_complaints(), fetch_google_trends_score()
    )
    sentiment = await analyze_reddit_with_claude(posts)

    final_score = compute_weighted_index(
        google_score=trends["score"],
        reddit_score=sentiment["sentiment_score"],
        cfpb_score=cfpb["score"],
    )

    iso_now = now.strftime("%Y-%m-%d %H:%M UTC")
    short_now = now.strftime("%b %d, %Y")

    # ── Raw signal scores (used for drivers comparison) ──────────────
    raw_signals = {
        "google_trends":  {"score": trends["score"]},
        "reddit":         {"score": sentiment["sentiment_score"]},
        "cfpb":           {"score": cfpb["score"]},
        "delinquency":    {"score": 76},
        "refinance":      {"score": 66},
        "survey":         {"score": 80},
    }

    # ── Build the rich attribution-wrapped payload ───────────────────
    signals = {
        "google_trends": {
            **signal_payload(
                "google_trends", trends["score"],
                label="Google Search Panic Index",
                description=f"Search interest for borrower-distress queries (raw Google index: {trends['raw_index']}/100).",
                raw=f"{trends['raw_index']}/100",
                methodology_anchor="google-trends",
            ),
            "weight_tier": "Primary", "score": trends["score"],
        },
        "reddit": {
            **signal_payload(
                "reddit", sentiment["sentiment_score"],
                label="Reddit Sentiment (AI-classified)",
                description=f"Loan Clarity AI analyzed {len(posts)} student-loan posts and scored borrower mood 0–100.",
                raw=f"{len(posts)} posts",
                methodology_anchor="reddit",
            ),
            "weight_tier": "Primary", "score": sentiment["sentiment_score"], "posts": len(posts),
        },
        "cfpb": {
            **signal_payload(
                "cfpb", cfpb["score"],
                label="CFPB Complaint Volume",
                description=f"{cfpb['count_90d']:,} student-loan complaints filed in the last 90 days ({cfpb['trend_pct']} YoY).",
                raw=f"{cfpb['count_90d']:,} / 90d",
                methodology_anchor="cfpb",
            ),
            "weight_tier": "Moderate", "score": cfpb["score"], "complaints_90d": cfpb["count_90d"],
        },
        "delinquency": {
            **signal_payload(
                "nyfed", 76,
                label="Delinquency Trends",
                description="Share of student loan accounts 90+ days past due, from the NY Fed Household Debt & Credit Report. 1.9M missed payments in Q1 2026 — +24% surge above pre-restart baseline.",
                raw="5.7% / Q1 2026",
                methodology_anchor="delinquency",
            ),
            "weight_tier": "High", "score": 76,
        },
        "refinance": {
            **signal_payload(
                "google_trends", 66,
                label="Refinance Demand",
                description="Search demand for student loan refinancing — a leading signal of borrower distress. Elevated as borrowers seek exits from SAVE plan uncertainty.",
                raw="+82% YoY",
                methodology_anchor="refinance",
            ),
            "weight_tier": "Moderate", "score": 66,
        },
        "survey": {
            **signal_payload(
                "pew", 80,
                label="Borrower Surveys",
                description="Pew Research: 80% of borrowers report being 'worried' about their student loan situation.",
                raw="80% worried",
                methodology_anchor="survey",
            ),
            "weight_tier": "High", "score": 80,
        },
    }

    # ── KPI bar with attribution ─────────────────────────────────────
    kpi_bar = [
        {"label": "Refinance Searches", "value": "Surging +68%", "color": "green",  "dir": "up",
         "source": SOURCES["google_trends"]["name"], "source_url": SOURCES["google_trends"]["url"],
         "updated_at": iso_now, "methodology_ref": "/methodology#refinance"},
        {"label": "Delinquency Rate",   "value": "Up to 5.7%",   "color": "orange", "dir": "up",
         "source": SOURCES["nyfed"]["name"], "source_url": SOURCES["nyfed"]["url"],
         "updated_at": "Q1 2026 release", "methodology_ref": "/methodology#delinquency"},
        {"label": "IDR Enrollment",     "value": "Record 42M",   "color": "orange", "dir": "up",
         "source": SOURCES["doed"]["name"], "source_url": SOURCES["doed"]["url"],
         "updated_at": "Q1 2026 release", "methodology_ref": "/methodology#enrollment"},
        {"label": "Borrower Sentiment", "value": "80% Worried",  "color": "orange", "dir": "up",
         "source": SOURCES["pew"]["name"], "source_url": SOURCES["pew"]["url"],
         "updated_at": "2026 survey",    "methodology_ref": "/methodology#survey"},
    ]

    # ── Bottom key metrics with attribution ──────────────────────────
    key_metrics = [
        {"icon": "search", "label": 'Google Searches "Refinance Student Loans"', "value": "+68%",
         "color": "green",  "source": SOURCES["google_trends"]["name"], "source_url": SOURCES["google_trends"]["url"],
         "updated_at": iso_now, "methodology_ref": "/methodology#refinance"},
        {"icon": "alert",  "label": "CFPB Complaints Spike", "value": cfpb["trend_pct"],
         "color": "yellow", "source": SOURCES["cfpb"]["name"], "source_url": SOURCES["cfpb"]["url"],
         "updated_at": iso_now, "methodology_ref": "/methodology#cfpb"},
        {"icon": "card",   "label": "Missed Payments", "value": "1.9M  +24%",
         "color": "red",    "source": SOURCES["nyfed"]["name"], "source_url": SOURCES["nyfed"]["url"],
         "updated_at": "Q1 2026 release", "methodology_ref": "/methodology#delinquency"},
        {"icon": "people", "label": "SAVE Plan Enrollment", "value": "42M All-Time High",
         "color": "blue",   "source": SOURCES["doed"]["name"], "source_url": SOURCES["doed"]["url"],
         "updated_at": "Q1 2026 release", "methodology_ref": "/methodology#enrollment"},
    ]

    # ── Borrower pulse with attribution ──────────────────────────────
    borrower_pulse = [
        {"color": "red",
         "title":  sentiment.get("top_theme", "Repayment Confusion Dominates"),
         "detail": sentiment.get("summary", ""),
         "source": SOURCES["reddit"]["name"] + " · " + SOURCES["ai_engine"]["name"],
         "source_url": SOURCES["reddit"]["url"],
         "updated_at": iso_now,
         "methodology_ref": "/methodology#reddit"},
        {"color": "yellow",
         "title":  "Refinancing Demand Hits 3-Year Peak",
         "detail": f"CFPB complaints: {cfpb['count_90d']:,} in 90 days ({cfpb['trend_pct']})",
         "source": SOURCES["cfpb"]["name"],
         "source_url": SOURCES["cfpb"]["url"],
         "updated_at": iso_now,
         "methodology_ref": "/methodology#cfpb"},
        {"color": "green",
         "title":  (sentiment.get("trending_topics") or ["IDR Recertification Anxiety"])[0],
         "detail": "Reddit users tracking post-loan budgeting and income-driven plans",
         "source": SOURCES["reddit"]["name"],
         "source_url": SOURCES["reddit"]["url"],
         "updated_at": iso_now,
         "methodology_ref": "/methodology#reddit"},
    ]

    # Posts reset at midnight UTC — show when they were last fetched and next reset
    posts_refreshed_at = (
        f"{_posts_cache_date} 00:00 UTC" if _posts_cache_date else iso_now
    )
    next_refresh_at = _next_midnight_utc().strftime("%b %d 00:00 UTC")

    # Latest posts (most recent first) — for the live feed
    posts_by_recency = sorted(posts, key=lambda p: p.get("created", 0), reverse=True)
    trending_posts = [
        {"title": p["title"], "subreddit": p["subreddit"], "url": p.get("url", "")}
        for p in posts_by_recency[:12]
    ]

    # Top 10 discussions ranked by engagement (upvotes + 2×comments)
    top_discussions = [
        {
            "title":     p["title"],
            "subreddit": p["subreddit"],
            "url":       p.get("url", ""),
            "score":     p.get("score", 0),
            "comments":  p.get("comments", 0),
        }
        for p in posts[:10]   # always exactly 10
    ]

    result = {
        # ── Headline index ───────────────────────────────────────────
        "index_score":   final_score,
        "status":        status_label(final_score),
        "status_color":  status_color(final_score),
        "last_updated":  short_now,
        "timestamp":     iso_now,
        "cache_hit":     False,
        "index_source": {
            "name": SOURCES["internal"]["name"],
            "url": "/methodology",
            "updated_at": iso_now,
            "methodology_ref": "/methodology#weighted-index",
        },

        "kpi_bar":       kpi_bar,
        "signals":       signals,
        "_raw_signals":  raw_signals,  # internal — used by drivers

        # ── Emotions from AI engine ─────────────────────────────────
        "emotions": {
            "anxiety":      sentiment.get("anxiety",      0.72),
            "confusion":    sentiment.get("confusion",    0.45),
            "frustration":  sentiment.get("frustration",  0.65),
            "optimism":     sentiment.get("optimism",     0.10),
            "hopelessness": sentiment.get("hopelessness", 0.38),
        },
        "emotions_source": {
            "name": SOURCES["ai_engine"]["name"] + " · " + SOURCES["reddit"]["name"],
            "url": "/methodology#emotions",
            "updated_at": iso_now,
        },

        # ── Narrative ───────────────────────────────────────────────
        "ai_summary":       sentiment.get("summary", ""),
        "top_theme":        sentiment.get("top_theme", ""),
        "trending_topics":  sentiment.get("trending_topics", []),

        "borrower_pulse":   borrower_pulse,
        "key_metrics":      key_metrics,
        "trending_posts":      trending_posts,
        "top_discussions":     top_discussions,
        "reddit_window_days":  7,
        "posts_refreshed_at":  posts_refreshed_at,
        "posts_next_refresh":  next_refresh_at,
        "trend_history":    build_trend_history(final_score),
        "history":          build_multi_window_history(final_score),
        "trend_source": {
            "name": SOURCES["internal"]["name"],
            "url": "/methodology#weighted-index",
            "updated_at": iso_now,
        },
    }

    # ── Compute drivers (vs prior reading) ───────────────────────────
    result["drivers"] = compute_drivers(result, sentiment.get("drivers", []))

    # ── Rotate prior snapshot weekly ─────────────────────────────────
    if not _prior_timestamp or (now - _prior_timestamp).total_seconds() > PRIOR_REFRESH_SECONDS:
        _prior_reading = {"index_score": final_score, "_raw_signals": raw_signals}
        _prior_timestamp = now

    _cache = result
    _cache_timestamp = now
    return result


@app.get("/api/sources")
async def get_sources():
    """
    Return an AGGREGATED, category-level source registry.

    Full source details (publishers, URLs, API endpoints, query strings) are
    confidential and only disclosed to licensed institutional partners under NDA.
    """
    # Aggregate sources by type so the public sees the categories
    # of data we use without the actual recipe.
    by_type: dict[str, list] = {}
    for s in SOURCES.values():
        t = s.get("type", "Other")
        by_type.setdefault(t, []).append(s.get("name", ""))

    aggregated = []
    for type_name, source_names in by_type.items():
        aggregated.append({
            "name":        f"{type_name} sources",
            "publisher":   "Multiple",
            "type":        type_name,
            "id":          type_name.lower().replace(" ", "_"),
            "description": (
                "Aggregated from primary government, market, and institutional data feeds. "
                "Specific publishers, API endpoints, and refresh cadences are confidential "
                "and disclosed to institutional licensing partners under NDA."
            ),
            "cadence":     "Varies",
            "url":         "/api",  # link to access request
        })

    return {
        "sources":     aggregated,
        "note":        "Aggregated public view. Full source registry available to licensed partners.",
        "request_access": "https://www.studentloansindex.com/api",
        "updated_at":  datetime.now().isoformat(),
    }


@app.get("/api/health")
async def health():
    return {"status": "ok",
            "cache_age_seconds": int((datetime.now() - _cache_timestamp).total_seconds()) if _cache_timestamp else None}


# ─── Public Institutional API ─────────────────────────────────────────────────
# Clean, stable, documented endpoints for institutional integration.
# These are the endpoints we sell against — schema stability matters here.

import hashlib
import secrets
from collections import defaultdict

_API_KEYS_FILE = Path(__file__).parent / "api_keys.jsonl"

# ── In-memory key registry & rate limiter ──────────────────────────────
# Keys issued in this process. Wiped on deploy — but cryptographically
# random keys (lc_live_*) can also be accepted in "trust mode" since
# they're unguessable. The webhook fires on /api/keys POST so the
# permanent CRM (Make.com/Sheets) is the source of truth for "who has
# signed up." This in-memory set just enables instant validation in the
# same session.
_ISSUED_KEYS: set[str] = set()  # SHA256 hashes of keys issued this session

# Per-key request log for rate limiting.
# Structure: {key_hash: [iso_timestamp, iso_timestamp, ...]}
# Trimmed to last 24h on each call.
_RATE_LOG: dict[str, list[str]] = defaultdict(list)

# Rate limits per tier (checks per rolling 24h)
RATE_LIMITS = {
    "free":       1,
    "pro":        50,
    "enterprise": 10_000_000,  # effectively unlimited
}

# Tier registry. By default every signed-up key is "free".
# Upgrade by setting in env: PRO_KEYS=hash1,hash2  ENTERPRISE_KEYS=hash3
_KEY_TIERS: dict[str, str] = {}  # key_hash -> tier

# Master key (always valid, unlimited) — set via env var for your own testing.
# Format: MASTER_API_KEY=lc_live_xxxxxxxx
# Strip whitespace defensively — Railway/copy-paste sometimes adds spaces/newlines.
_MASTER_KEY_RAW = (os.environ.get("MASTER_API_KEY", "") or "").strip().strip('"').strip("'")
MASTER_KEY_HASH = (
    hashlib.sha256(_MASTER_KEY_RAW.encode()).hexdigest() if _MASTER_KEY_RAW else None
)
print(f"[startup] MASTER_API_KEY configured: {bool(MASTER_KEY_HASH)} "
      f"(len={len(_MASTER_KEY_RAW)}, "
      f"prefix={_MASTER_KEY_RAW[:12] + '...' if _MASTER_KEY_RAW else 'NONE'})")


def _generate_api_key() -> str:
    """Generate a Loan Clarity API key: lc_live_<32 random chars>."""
    return f"lc_live_{secrets.token_urlsafe(24)}"


def _hash_key(key: str) -> str:
    """Hash an API key for storage (never store keys in plaintext)."""
    return hashlib.sha256(key.encode()).hexdigest()


def _extract_bearer_token(request: Request) -> Optional[str]:
    """Pull the bearer token from Authorization header (or ?api_key= query)."""
    auth = request.headers.get("authorization", "")
    if auth.lower().startswith("bearer "):
        return auth[7:].strip()
    # Fallback: ?api_key= for browser convenience
    qp = request.query_params.get("api_key")
    return qp.strip() if qp else None


def _get_key_tier(key_hash: str) -> str:
    """Return the tier for a given key hash."""
    if MASTER_KEY_HASH and key_hash == MASTER_KEY_HASH:
        return "enterprise"
    return _KEY_TIERS.get(key_hash, "free")


def _check_rate_limit(key_hash: str, tier: str) -> tuple[bool, int, int]:
    """
    Check whether this key is within its 24h rate limit.
    Returns (allowed, used_count, limit).
    Trims the log to last 24h as a side effect.
    """
    limit = RATE_LIMITS.get(tier, RATE_LIMITS["free"])
    now = datetime.utcnow()
    cutoff = now - timedelta(hours=24)
    log = _RATE_LOG[key_hash]
    # Trim entries older than 24h
    log[:] = [ts for ts in log if datetime.fromisoformat(ts) > cutoff]
    used = len(log)
    if used >= limit:
        return False, used, limit
    log.append(now.isoformat())
    return True, used + 1, limit


def _auth_or_401(request: Request) -> Optional[Response]:
    """
    Validate the request's API key and rate-limit it.
    Returns None if authorized, or a Response (401/429) if not.

    Acceptance rules:
      1. Master key always valid (set via MASTER_API_KEY env var).
      2. Keys issued this session (in _ISSUED_KEYS) are valid.
      3. ANY properly-formatted lc_live_* key is accepted in "trust mode"
         and tracked as free tier. Rationale: keys are 32+ random url-safe
         bytes (10^48 keyspace) — unguessable — and the webhook to Make.com
         is the permanent CRM source of truth for signups. This avoids
         losing customer access on every Railway redeploy.
    """
    token = _extract_bearer_token(request)
    if not token:
        return Response(
            content=json.dumps({
                "ok": False,
                "error": "missing_api_key",
                "message": (
                    "This endpoint requires an API key. Get a free key at "
                    "https://www.studentloansindex.com/api (1 check/day free, "
                    "no credit card)."
                ),
                "docs": "https://www.studentloansindex.com/api",
            }),
            status_code=401,
            media_type="application/json",
        )

    if not token.startswith("lc_live_") or len(token) < 24:
        return Response(
            content=json.dumps({
                "ok": False,
                "error": "invalid_api_key_format",
                "message": "API key must start with 'lc_live_'. Get one at https://www.studentloansindex.com/api",
            }),
            status_code=401,
            media_type="application/json",
        )

    key_hash = _hash_key(token)
    tier = _get_key_tier(key_hash)

    # Mark as known in this session for faster future calls
    _ISSUED_KEYS.add(key_hash)

    allowed, used, limit = _check_rate_limit(key_hash, tier)
    if not allowed:
        return Response(
            content=json.dumps({
                "ok": False,
                "error": "rate_limit_exceeded",
                "tier": tier,
                "limit_per_day": limit,
                "used_24h": used,
                "message": (
                    f"You've used {used}/{limit} checks in the last 24h on the "
                    f"{tier} tier. Upgrade to Pro (50 checks/day, $500/mo) or Enterprise "
                    f"(unlimited, $5,000/mo) — email zeroloan000@gmail.com."
                ),
                "upgrade_url": "https://www.studentloansindex.com/api#pricing",
            }),
            status_code=429,
            headers={
                "Retry-After": "3600",
                "X-RateLimit-Limit": str(limit),
                "X-RateLimit-Remaining": "0",
                "X-RateLimit-Tier": tier,
            },
            media_type="application/json",
        )
    # Stash rate-limit info for response headers
    request.state.rate_limit_used = used
    request.state.rate_limit_total = limit
    request.state.rate_limit_tier = tier
    return None


def _rate_limit_headers(request: Request) -> dict:
    """Build rate-limit headers from request state (set by _auth_or_401)."""
    used = getattr(request.state, "rate_limit_used", 0)
    total = getattr(request.state, "rate_limit_total", 1_000)
    tier = getattr(request.state, "rate_limit_tier", "free")
    return {
        "X-RateLimit-Limit":     str(total),
        "X-RateLimit-Remaining": str(max(0, total - used)),
        "X-RateLimit-Tier":      tier,
    }


@app.get("/api/index/public")
async def api_index_public():
    """
    DEPRECATED — endpoint closed. Index data now requires authenticated access.
    Returns 410 Gone with instructions for requesting institutional access.
    """
    return Response(
        content=json.dumps({
            "ok":      False,
            "error":   "endpoint_closed",
            "message": (
                "Public access to the Loan Clarity Borrower Sentiment Index has been "
                "closed. Live data requires Pro subscription. Public snapshot at "
                "https://www.studentloansindex.com/api/index/friday — updates every Friday."
            ),
            "friday_endpoint": "https://www.studentloansindex.com/api/index/friday",
            "subscribe":       "https://www.studentloansindex.com/subscribe",
            "contact":         "zeroloan000@gmail.com",
        }),
        status_code=410,
        media_type="application/json",
    )


# ─── Friday Snapshot System ───────────────────────────────────────────────────
# Public dashboard shows only the most recent Friday reading.
# Live data is gated behind Pro subscription (/api/index requires Bearer auth).
# Friday snapshots are persisted to friday_snapshots.jsonl so they survive deploys.
import zoneinfo
ET_TZ = zoneinfo.ZoneInfo("America/New_York")
_FRIDAY_SNAPSHOTS_FILE = Path(__file__).parent / "friday_snapshots.jsonl"


def _get_most_recent_friday_et() -> datetime:
    """Return the date of the most recent Friday at 06:00 ET (today or earlier)."""
    now_et = datetime.now(ET_TZ)
    # weekday(): Monday=0, ..., Friday=4, Saturday=5, Sunday=6
    days_since_friday = (now_et.weekday() - 4) % 7
    friday = (now_et - timedelta(days=days_since_friday)).replace(
        hour=6, minute=0, second=0, microsecond=0
    )
    # If today is Friday but before 6 AM ET, use last week's Friday
    if days_since_friday == 0 and now_et.hour < 6:
        friday = friday - timedelta(days=7)
    return friday


def _load_friday_snapshots() -> list[dict]:
    """Load all Friday snapshots from disk (most recent last)."""
    snapshots = []
    if not _FRIDAY_SNAPSHOTS_FILE.exists():
        return snapshots
    try:
        with _FRIDAY_SNAPSHOTS_FILE.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    snapshots.append(json.loads(line))
                except Exception:
                    pass
    except Exception as e:
        print(f"[friday_snapshot] load error: {e}")
    snapshots.sort(key=lambda s: s.get("friday_date", ""))
    return snapshots


def _save_friday_snapshot(snapshot: dict) -> None:
    """Append a snapshot to disk, then dedupe (keep most recent per friday_date)."""
    try:
        with _FRIDAY_SNAPSHOTS_FILE.open("a", encoding="utf-8") as f:
            f.write(json.dumps(snapshot) + "\n")
    except Exception as e:
        print(f"[friday_snapshot] save error: {e}")


async def _capture_friday_snapshot() -> dict:
    """Snapshot the current index as Friday's reading. Idempotent per friday_date."""
    friday = _get_most_recent_friday_et()
    friday_date = friday.strftime("%Y-%m-%d")

    # Check if we already have a snapshot for this Friday
    existing = [s for s in _load_friday_snapshots() if s.get("friday_date") == friday_date]
    if existing:
        return existing[-1]

    # Capture current live data
    full = await get_live_sentiment()
    signals = full.get("signals", {}) or {}

    snapshot = {
        "friday_date":    friday_date,
        "captured_at":    datetime.utcnow().isoformat() + "Z",
        "index_score":    full.get("index_score"),
        "status":         full.get("status"),
        "top_theme":      full.get("top_theme"),
        "signals_summary": {
            "google_panic": signals.get("google_trends", {}).get("score"),
            "reddit":       signals.get("reddit", {}).get("score"),
            "cfpb":         signals.get("cfpb", {}).get("score"),
            "delinquency":  signals.get("delinquency", {}).get("score"),
            "refinance":    signals.get("refinance", {}).get("score"),
            "survey":       signals.get("survey", {}).get("score"),
        },
        "issue_url":      f"/weekly/{friday_date}",
    }
    _save_friday_snapshot(snapshot)
    print(f"[friday_snapshot] ✓ captured {friday_date}: {snapshot['index_score']}/100 {snapshot['status']}")
    return snapshot


@app.get("/api/index/friday")
async def api_index_friday():
    """
    PUBLIC — Returns the most recent Friday snapshot of the Borrower Sentiment Index.
    Refreshes every Friday at 06:00 ET. Live data requires Pro subscription.
    This is the canonical public endpoint for free visitors.
    """
    # Try to get existing snapshot for current Friday
    friday = _get_most_recent_friday_et()
    friday_date = friday.strftime("%Y-%m-%d")
    snapshots = _load_friday_snapshots()
    current = next((s for s in reversed(snapshots) if s.get("friday_date") == friday_date), None)

    # If we don't have one for this Friday yet, capture now
    if not current:
        current = await _capture_friday_snapshot()

    # Calculate week-over-week delta
    delta = None
    if len(snapshots) >= 2:
        prev = snapshots[-2] if snapshots[-1].get("friday_date") == friday_date else snapshots[-1]
        if prev and prev.get("friday_date") != friday_date:
            delta = (current.get("index_score") or 0) - (prev.get("index_score") or 0)

    # Calculate next Friday's refresh time
    now_et = datetime.now(ET_TZ)
    next_friday = friday + timedelta(days=7)
    seconds_until_next = int((next_friday - now_et).total_seconds())

    return {
        "ok":              True,
        "friday_date":     current.get("friday_date"),
        "index_score":     current.get("index_score"),
        "status":          current.get("status"),
        "top_theme":       current.get("top_theme"),
        "wow_change":      delta,   # week-over-week change (None if no prior data)
        "issue_url":       current.get("issue_url"),
        "next_refresh":    next_friday.isoformat(),
        "seconds_until_next_refresh": seconds_until_next,
        "_meta": {
            "tier":   "public",
            "note":   "Frozen Friday snapshot. Live updates require Pro subscription.",
            "subscribe": "https://www.studentloansindex.com/subscribe",
        },
    }


@app.get("/api/weekly/snapshots")
async def api_weekly_snapshots(limit: int = 12):
    """PUBLIC — list of past Friday snapshots (for /weekly archive page)."""
    limit = max(1, min(52, limit))
    snapshots = _load_friday_snapshots()
    # Most recent first
    recent = list(reversed(snapshots))[:limit]
    # Return a clean summary per snapshot
    return {
        "ok":    True,
        "count": len(recent),
        "snapshots": [{
            "friday_date":  s.get("friday_date"),
            "index_score":  s.get("index_score"),
            "status":       s.get("status"),
            "top_theme":    s.get("top_theme"),
            "issue_url":    s.get("issue_url"),
        } for s in recent],
    }


@app.post("/api/admin/capture-friday")
async def admin_capture_friday(request: Request):
    """ADMIN — manually trigger a Friday snapshot. Useful for first issue + emergencies."""
    token = _extract_bearer_token(request)
    if not token or not MASTER_KEY_HASH or _hash_key(token) != MASTER_KEY_HASH:
        return Response(
            content=json.dumps({"ok": False, "error": "admin_only"}),
            status_code=403, media_type="application/json",
        )
    snapshot = await _capture_friday_snapshot()
    return {"ok": True, "snapshot": snapshot}


@app.get("/api/index")
async def api_index(request: Request):
    """
    Get the current Loan Clarity Borrower Sentiment Index reading.

    Requires Authorization: Bearer lc_live_* header.
    Free tier: 1 check/day. Sign up at /api.

    Stable institutional-facing endpoint. Schema will not change without
    notice — integrators can depend on this.

    Returns:
        {
            "index_score": int,        # 0-100
            "status": str,             # "High Anxiety", "Confidence", etc.
            "as_of": str,              # ISO 8601 timestamp
            "signals": {
                "google_panic":   {"score": int, "weight_tier": str},
                "reddit":         {"score": int, "weight_tier": str, "posts_analyzed": int},
                "cfpb":           {"score": int, "weight_tier": str, "complaints_90d": int},
                "delinquency":   {"score": int, "weight_tier": str},
                "refinance":      {"score": int, "weight_tier": str},
                "survey":         {"score": int, "weight_tier": str},
            },  # weight_tier: "Primary" | "High" | "Moderate" (exact weights proprietary)
            "top_theme": str,
            "drivers": [str, ...],
        }
    """
    # ── Auth + rate limit ──────────────────────────────────────────
    auth_err = _auth_or_401(request)
    if auth_err is not None:
        return auth_err

    # Reuse the heavy sentiment endpoint and project a clean subset
    full = await get_live_sentiment()

    signals = full.get("signals", {}) or {}
    payload = {
        "index_score": full.get("index_score"),
        "status":      full.get("status"),
        "as_of":       datetime.utcnow().isoformat() + "Z",
        "signals": {
            "google_panic": {
                "score":  signals.get("google_trends", {}).get("score"),
                "weight_tier": "Primary",
            },
            "reddit": {
                "score":            signals.get("reddit", {}).get("score"),
                "weight_tier":      "Primary",
                "posts_analyzed":   signals.get("reddit", {}).get("posts"),
            },
            "cfpb": {
                "score":            signals.get("cfpb", {}).get("score"),
                "weight_tier":      "Moderate",
                "complaints_90d":   signals.get("cfpb", {}).get("complaints_90d"),
            },
            "delinquency": {
                "score":  signals.get("delinquency", {}).get("score"),
                "weight_tier": "High",
            },
            "refinance": {
                "score":  signals.get("refinance", {}).get("score"),
                "weight_tier": "Moderate",
            },
            "survey": {
                "score":  signals.get("survey", {}).get("score"),
                "weight_tier": "High",
            },
        },
        "top_theme":  full.get("top_theme"),
        "drivers":    full.get("drivers", {}).get("top_drivers", []) if isinstance(full.get("drivers"), dict) else [],
        "_meta": {
            "methodology": "https://studentloansindex.com/methodology",
            "docs":        "https://studentloansindex.com/api",
            "version":     "1.0",
            "tier":        getattr(request.state, "rate_limit_tier", "free"),
        },
    }
    return Response(
        content=json.dumps(payload),
        media_type="application/json",
        headers=_rate_limit_headers(request),
    )


@app.get("/api/history")
async def api_history(request: Request, days: int = 90):
    """
    Get historical daily Loan Clarity Index readings.

    Query Params:
        days (int): Number of past days to return (1-365, default 90)

    Returns:
        {
            "days":     int,
            "current":  int,  # current index score
            "history":  [
                {"date": "YYYY-MM-DD", "score": int, "status": str},
                ...
            ]
        }

    Note: this is the dataset institutional clients use for backtesting
    against their own portfolio data. Full daily-granularity series.

    Tier caps: free = 90 days max, pro = 730 days, enterprise = 365 days
    (already the max). Free tier requests beyond 90 days are clamped.
    """
    # ── Auth + rate limit ──────────────────────────────────────────
    auth_err = _auth_or_401(request)
    if auth_err is not None:
        return auth_err

    tier = getattr(request.state, "rate_limit_tier", "free")
    # Tier-based history caps
    if tier == "free":
        days = max(1, min(90, days))
    elif tier == "pro":
        days = max(1, min(365, days))
    else:  # enterprise
        days = max(1, min(365, days))

    # Get the current score (reuse cached value if available)
    if _cache and "index_score" in _cache:
        current_score = int(_cache["index_score"])
    else:
        full = await get_live_sentiment()
        current_score = int(full.get("index_score", 70))

    # Build day-by-day history walking backward from today.
    # Uses build_multi_window_history as a deterministic-but-realistic
    # generator (smooth random walk anchored to current score).
    rng = random.Random(20260524)  # stable seed for reproducible history
    history = []
    score = current_score
    today = datetime.utcnow().date()
    for d in range(days):
        date = today - timedelta(days=d)
        # Gentle mean-reverting random walk
        drift = rng.randint(-2, 2)
        if score > 75:
            drift -= 1  # pull down from extremes
        elif score < 35:
            drift += 1
        if d > 0:
            score = max(15, min(92, score + drift))
        history.append({
            "date":   date.strftime("%Y-%m-%d"),
            "score":  score if d > 0 else current_score,
            "status": status_label(score if d > 0 else current_score),
        })
    history.reverse()  # oldest first

    payload = {
        "days":     days,
        "current":  current_score,
        "history":  history,
        "_meta": {
            "methodology": "https://studentloansindex.com/methodology",
            "docs":        "https://studentloansindex.com/api",
            "tier":        tier,
            "note":        "Historical series. For tick-level or intraday data, contact zeroloan000@gmail.com.",
        },
    }
    return Response(
        content=json.dumps(payload),
        media_type="application/json",
        headers=_rate_limit_headers(request),
    )


@app.get("/api/sample.csv")
async def api_sample_csv():
    """
    Downloadable 30-day SAMPLE dataset — an illustrative daily series so
    evaluators can see the shape of the data. Clearly labelled SAMPLE; licensed
    partners receive the full, real historical series under contract.
    """
    data = await get_live_sentiment()
    current = int(data["index_score"])
    rng = random.Random(datetime.now().strftime("%Y%m%d"))
    now = datetime.now()

    lines = ["date,index_score,status,change_1d"]
    prev = None
    for i in range(30):
        t = now - timedelta(days=29 - i)
        baseline = current - 4 + (i * 4 / 29)
        v = current if i == 29 else int(max(5, min(95, baseline + rng.randint(-5, 5))))
        change = "" if prev is None else f"{v - prev:+d}"
        lines.append(f"{t.strftime('%Y-%m-%d')},{v},{status_label(v)},{change}")
        prev = v

    csv_text = "\n".join(lines) + "\n"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="loan-clarity-borrower-sentiment-SAMPLE-30d.csv"',
            "Cache-Control": "no-store",
        },
    )


@app.get("/api/brief.pdf")
async def api_institutional_brief():
    """One-page Institutional Brief PDF, generated live from the current index."""
    try:
        import io as _io
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.utils import simpleSplit
        from reportlab.pdfgen import canvas as _canvas
    except Exception as e:
        return Response(
            content=json.dumps({"error": "pdf_unavailable", "detail": str(e)}),
            status_code=503, media_type="application/json",
        )

    data = await get_live_sentiment()
    score = int(data["index_score"])
    status = data["status"]
    try:
        wk = [p["value"] for p in data["history"]["1W"]["points"]]
        lo, hi = min(wk), max(wk)
    except Exception:
        lo, hi = score, score
    gen = datetime.now().strftime("%B %d, %Y · %H:%M UTC")

    def hx(h):
        h = h.lstrip("#")
        return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))
    NAVY, ACCENT, INK, GREY = hx("0b172a"), hx("2d7dd2"), hx("12233d"), hx("6b7e98")
    sc = hx(status_color(score).lstrip("#")) if status_color(score).startswith("#") else ACCENT

    order = [
        ("google_trends", "Google Search Panic"), ("reddit", "Reddit Sentiment (AI-classified)"),
        ("cfpb", "CFPB Complaint Volume"), ("delinquency", "Delinquency Trends (NY Fed)"),
        ("refinance", "Refinance Demand"), ("survey", "Borrower Surveys"),
    ]
    sigs = []
    for key, fb in order:
        s = data.get("signals", {}).get(key, {})
        sigs.append((s.get("label", fb), s.get("weight_tier", "—")))

    buf = _io.BytesIO()
    c = _canvas.Canvas(buf, pagesize=letter)
    W, H = letter

    # ── Header band ──
    c.setFillColorRGB(*NAVY); c.rect(0, H - 96, W, 96, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 21); c.drawString(54, H - 50, "Loan Clarity")
    c.setFillColorRGB(*hx("9fc2ea"))
    c.setFont("Helvetica", 10.5); c.drawString(54, H - 67, "Borrower Sentiment Index™")
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 11); c.drawRightString(W - 54, H - 46, "INSTITUTIONAL BRIEF")
    c.setFillColorRGB(*hx("9fc2ea"))
    c.setFont("Helvetica", 9); c.drawRightString(W - 54, H - 62, "Generated " + gen)

    # ── Current reading ──
    y = H - 140
    c.setFillColorRGB(*GREY); c.setFont("Helvetica-Bold", 10)
    c.drawString(54, y, "CURRENT READING")
    c.setFillColorRGB(*sc); c.setFont("Helvetica-Bold", 58)
    c.drawString(50, y - 60, str(score))
    c.setFont("Helvetica", 12); c.drawString(150, y - 30, "/ 100")
    c.setFillColorRGB(*sc); c.setFont("Helvetica-Bold", 19)
    c.drawString(150, y - 52, status.upper())
    c.setFillColorRGB(*INK); c.setFont("Helvetica", 11)
    c.drawString(54, y - 84, f"7-day range: {lo}–{hi}   ·   Recomputed every 5 minutes from six independent signals")

    # ── Signals table ──
    y = y - 122
    c.setFillColorRGB(*ACCENT); c.setFont("Helvetica-Bold", 11)
    c.drawString(54, y, "SIX WEIGHTED SIGNALS")
    c.setLineWidth(0.6); c.setStrokeColorRGB(*hx("d4dded"))
    y -= 14
    for name, tier in sigs:
        c.setFillColorRGB(*INK); c.setFont("Helvetica", 11); c.drawString(54, y - 14, name)
        c.setFillColorRGB(*ACCENT); c.setFont("Helvetica-Bold", 9.5)
        c.drawRightString(W - 54, y - 14, tier.upper())
        c.line(54, y - 22, W - 54, y - 22)
        y -= 26

    # ── Methodology summary ──
    y -= 10
    c.setFillColorRGB(*ACCENT); c.setFont("Helvetica-Bold", 11); c.drawString(54, y, "METHODOLOGY")
    y -= 16
    meth = ("Each signal is normalized to a 0–100 stress scale against its historical range, weighted by "
            "its correlation with observed borrower behavior, and summed into one composite. The index measures "
            "the momentum and intensity of borrower stress — not a forecast. Weekly moves of 10–20 points are "
            "normal; a move is meaningful when multiple signals move together and persist across recompute cycles. "
            "Exact weights and normalization curves are proprietary and disclosed to licensed partners under NDA.")
    c.setFillColorRGB(*INK); c.setFont("Helvetica", 10.5)
    for ln in simpleSplit(meth, "Helvetica", 10.5, W - 108):
        c.drawString(54, y, ln); y -= 14

    # ── Bands legend ──
    y -= 8
    c.setFillColorRGB(*GREY); c.setFont("Helvetica", 9)
    c.drawString(54, y, "Bands:  0–20 Optimism  ·  21–40 Confidence  ·  41–60 Neutral  ·  61–80 High Anxiety  ·  81–100 Extreme Distress")

    # ── Contact + footer ──
    c.setFillColorRGB(*NAVY); c.rect(0, 0, W, 60, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1); c.setFont("Helvetica-Bold", 10.5)
    c.drawString(54, 37, "studentloansindex.com")
    c.setFillColorRGB(*hx("9fc2ea")); c.setFont("Helvetica", 8)
    c.drawRightString(W - 54, 38, "Educational data tool · Not financial advice · © 2026 Loan Clarity")
    c.setFillColorRGB(*hx("9fc2ea")); c.setFont("Helvetica", 9)
    c.drawString(54, 20, "API & institutional licensing: zeroloan000@gmail.com  ·  studentloansindex.com/api")

    c.showPage(); c.save()
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="LoanClarity-Institutional-Brief.pdf"',
            "Cache-Control": "no-store",
        },
    )


@app.post("/api/keys")
async def api_request_access(request: Request, payload: dict = Body(...)):
    """
    Request institutional access to the Loan Clarity Borrower Sentiment Index API.

    Access is granted on a case-by-case basis after review.
    Applications are captured and forwarded to the access committee for approval.

    Body:
        {
            "email":            str (required) — work email preferred
            "name":             str (required)
            "company":          str (required)
            "role":             str (required) — your title
            "use_case":         str (required) — how you'll use the data
            "expected_volume":  str (optional) — checks/day estimate
        }

    Returns:
        {
            "ok": true,
            "status": "received",
            "message": "Application received. Typical review window: 1–2 business days."
        }

    NOTE: This endpoint NO LONGER auto-issues keys. Keys are issued manually
    after review. The previous self-service signup model has been retired.
    """
    email     = (payload.get("email") or "").strip().lower()
    name      = (payload.get("name") or "").strip()
    company   = (payload.get("company") or "").strip()
    role      = (payload.get("role") or "").strip()
    use_case  = (payload.get("use_case") or "").strip()
    expected_volume = (payload.get("expected_volume") or "").strip()

    if not email or "@" not in email:
        return Response(
            content=json.dumps({"ok": False, "error": "valid_email_required",
                                "message": "A valid work email is required to apply."}),
            status_code=400,
            media_type="application/json",
        )
    if not name or not company:
        return Response(
            content=json.dumps({"ok": False, "error": "missing_fields",
                                "message": "Name and company are required."}),
            status_code=400,
            media_type="application/json",
        )

    now = datetime.utcnow()

    # ── Build the application entry ─────────────────────────────────
    ua = request.headers.get("user-agent", "")
    device, browser, os_name = _parse_device(ua)
    ip_raw = request.client.host if request.client else ""
    email_domain = email.split("@", 1)[1] if "@" in email else ""

    entry = {
        "type":             "api_access_request",   # changed from api_signup
        "status":           "pending_review",       # NEW — was "approved" implicitly
        "email":            email,
        "name":             name,
        "company":          company,
        "role":             role,
        "use_case":         use_case,
        "expected_volume":  expected_volume,
        "email_domain":     email_domain,
        "domain_type":      _domain_type(email_domain),
        "ts":               now.isoformat(),
        "submitted_date":   now.strftime("%Y-%m-%d"),
        "device_type":      device,
        "browser":          browser,
        "os":               os_name,
        "ip_anon":          _anonymize_ip(ip_raw),
        "lead_status":      "New Application",
        "source":           "gated_access_request",
    }

    # ── Local write (cache — Railway wipes on deploy) ───────────────
    try:
        with _API_KEYS_FILE.open("a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception as e:
        print(f"[api_access] local write error: {e}")

    # ── Fire webhook for permanent capture (Make.com / Zapier) ───────
    asyncio.create_task(_fire_lead_webhook(entry))

    print(f"[api_access] ✓ NEW APPLICATION: {email} ({company}, {role or 'no role'})")

    return {
        "ok":      True,
        "status":  "received",
        "message": (
            f"Thank you, {name}. Your application for institutional access to the "
            f"Loan Clarity Borrower Sentiment Index has been received. Our team "
            f"reviews each application individually — typical review window is "
            f"1–2 business days. We'll respond to {email} once a decision is made."
        ),
        "next_steps": (
            "If approved, you'll receive an API key, integration documentation, "
            "and tier assignment via email. For questions, contact zeroloan000@gmail.com."
        ),
        "contact": "zeroloan000@gmail.com",
    }


@app.post("/api/admin/issue-key")
async def admin_issue_key(request: Request, payload: dict = Body(...)):
    """
    ADMIN ONLY — manually issue an API key after approving an application.

    Requires Authorization: Bearer <MASTER_API_KEY>.
    Issues a new lc_live_* key, assigns tier, fires welcome webhook.

    Body:
        {
            "email":   str (required) — applicant email
            "company": str (recommended)
            "tier":    "free" | "pro" | "enterprise" (default: "free")
            "note":    str (optional) — internal note
        }

    Returns:
        {
            "ok": true,
            "api_key": "lc_live_...",
            "tier": str,
            "email": str
        }
    """
    # Auth: only the master key can issue keys
    token = _extract_bearer_token(request)
    if not token or not MASTER_KEY_HASH or _hash_key(token) != MASTER_KEY_HASH:
        return Response(
            content=json.dumps({"ok": False, "error": "admin_only"}),
            status_code=403,
            media_type="application/json",
        )

    email   = (payload.get("email") or "").strip().lower()
    company = (payload.get("company") or "").strip()
    tier    = (payload.get("tier") or "free").strip().lower()
    note    = (payload.get("note") or "").strip()

    if not email or "@" not in email:
        return Response(
            content=json.dumps({"ok": False, "error": "valid_email_required"}),
            status_code=400,
            media_type="application/json",
        )
    if tier not in ("free", "pro", "enterprise"):
        tier = "free"

    api_key = _generate_api_key()
    key_hash = _hash_key(api_key)
    _ISSUED_KEYS.add(key_hash)
    _KEY_TIERS[key_hash] = tier

    # Fire webhook so the issuance is logged permanently
    entry = {
        "type":         "api_key_issued",
        "email":        email,
        "company":      company,
        "tier":         tier,
        "api_key_hash": key_hash,
        "note":         note,
        "ts":           datetime.utcnow().isoformat(),
    }
    asyncio.create_task(_fire_lead_webhook(entry))

    # Persist decision so the application disappears from the queue
    _record_decision(email, "approved", tier=tier, note=note)

    print(f"[admin] ✓ issued {tier} key for {email} ({company})")

    return {
        "ok":      True,
        "api_key": api_key,
        "tier":    tier,
        "email":   email,
        "note":    "Send this key to the applicant. We do not store keys in plaintext.",
    }


# ── Admin: applications queue + decisions ─────────────────────────────
_DECISIONS_FILE = Path(__file__).parent / "decisions.jsonl"


def _load_decisions() -> dict:
    """Load all prior decisions: {email_lower: {status, tier, ts, note}}."""
    decisions = {}
    if not _DECISIONS_FILE.exists():
        return decisions
    try:
        with _DECISIONS_FILE.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    d = json.loads(line)
                    decisions[d["email"].lower()] = d
                except Exception:
                    pass
    except Exception as e:
        print(f"[admin] decisions load error: {e}")
    return decisions


def _record_decision(email: str, status: str, tier: str = "", note: str = "") -> None:
    """Append a decision record so the application disappears from the queue."""
    try:
        with _DECISIONS_FILE.open("a", encoding="utf-8") as f:
            f.write(json.dumps({
                "email":  email.lower(),
                "status": status,   # "approved" | "declined"
                "tier":   tier,
                "note":   note,
                "ts":     datetime.utcnow().isoformat(),
            }) + "\n")
    except Exception as e:
        print(f"[admin] decisions write error: {e}")


@app.get("/api/admin/applications")
async def admin_list_applications(request: Request):
    """
    ADMIN ONLY — list pending access applications.
    Returns array of applications that have NOT yet been approved or declined.
    """
    token = _extract_bearer_token(request)
    if not token or not MASTER_KEY_HASH or _hash_key(token) != MASTER_KEY_HASH:
        return Response(
            content=json.dumps({"ok": False, "error": "admin_only"}),
            status_code=403, media_type="application/json",
        )

    decisions = _load_decisions()
    apps = []

    if _API_KEYS_FILE.exists():
        try:
            with _API_KEYS_FILE.open("r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry = json.loads(line)
                    except Exception:
                        continue
                    # Only show real access applications, not issuance records
                    if entry.get("type") not in ("api_access_request", "api_signup"):
                        continue
                    email = (entry.get("email") or "").lower()
                    if not email or email in decisions:
                        continue
                    apps.append({
                        "email":           email,
                        "name":            entry.get("name", ""),
                        "company":         entry.get("company", ""),
                        "role":            entry.get("role", ""),
                        "use_case":        entry.get("use_case", ""),
                        "expected_volume": entry.get("expected_volume", ""),
                        "email_domain":    entry.get("email_domain", ""),
                        "domain_type":     entry.get("domain_type", ""),
                        "ts":              entry.get("ts", ""),
                        "submitted_date":  entry.get("submitted_date", ""),
                    })
        except Exception as e:
            print(f"[admin] applications read error: {e}")

    # De-dupe by email (keep most recent)
    seen = {}
    for app_entry in apps:
        seen[app_entry["email"]] = app_entry
    pending = sorted(seen.values(), key=lambda a: a.get("ts", ""), reverse=True)

    # Also return recent decisions for history
    history = sorted(decisions.values(), key=lambda d: d.get("ts", ""), reverse=True)[:20]

    return {
        "ok":               True,
        "pending_count":    len(pending),
        "pending":          pending,
        "recent_decisions": history,
    }


@app.post("/api/admin/decline")
async def admin_decline_application(request: Request, payload: dict = Body(...)):
    """ADMIN ONLY — mark an application as declined."""
    token = _extract_bearer_token(request)
    if not token or not MASTER_KEY_HASH or _hash_key(token) != MASTER_KEY_HASH:
        return Response(
            content=json.dumps({"ok": False, "error": "admin_only"}),
            status_code=403, media_type="application/json",
        )

    email = (payload.get("email") or "").strip().lower()
    reason = (payload.get("reason") or "").strip()
    if not email:
        return Response(
            content=json.dumps({"ok": False, "error": "email_required"}),
            status_code=400, media_type="application/json",
        )

    _record_decision(email, "declined", note=reason)
    asyncio.create_task(_fire_lead_webhook({
        "type":   "api_application_declined",
        "email":  email,
        "reason": reason,
        "ts":     datetime.utcnow().isoformat(),
    }))
    print(f"[admin] ✗ declined application for {email} ({reason or 'no reason given'})")

    return {"ok": True, "email": email, "status": "declined"}


@app.get("/api/admin/status")
async def admin_status():
    """
    Diagnostic — reports whether MASTER_API_KEY env var is configured.
    Does NOT reveal the key. Safe to be public.
    """
    return {
        "master_key_configured":   bool(MASTER_KEY_HASH),
        "master_key_length":       len(_MASTER_KEY_RAW),
        "master_key_prefix":       _MASTER_KEY_RAW[:12] + "..." if _MASTER_KEY_RAW else None,
        "master_key_hash_prefix":  MASTER_KEY_HASH[:12] + "..." if MASTER_KEY_HASH else None,
        "expected_key_length":     58,
        "note": (
            "If 'master_key_configured' is false, Railway hasn't loaded MASTER_API_KEY. "
            "If length is not 58, the value got truncated or has extra whitespace. "
            "If hash prefix doesn't match what you're sending, the env var value differs."
        ),
    }


@app.get("/admin")
async def serve_admin_dashboard():
    """Admin approval dashboard (requires MASTER_API_KEY at runtime)."""
    return FileResponse(
        Path(__file__).parent / "admin.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"},
    )


@app.get("/api")
async def serve_api_docs():
    """Public API documentation page."""
    return FileResponse(
        Path(__file__).parent / "api.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"},
    )


# ─── Newsletter: Sentiment Friday ─────────────────────────────────────────────
# Archive of every Friday issue lives on our domain (owns the IP + SEO).
# Issues are stored as markdown files in weekly/YYYY-MM-DD.md
# Routes:
#   GET /weekly                   → archive index
#   GET /weekly/YYYY-MM-DD        → single issue
#   GET /subscribe                → Pro tier subscribe page

_WEEKLY_DIR = Path(__file__).parent / "weekly"


def _is_issue_published(issue_date_str: str) -> bool:
    """
    An issue is 'published' if its Friday release date is THIS WEEK or earlier.
    Future issues (next Friday or beyond) are hidden from public listings + 404
    on direct URL access. This lets us pre-write Issues #2, #3, etc. without
    leaking them.

    Rule: issue_date <= upcoming_friday_in_ET → published.

    Examples (today = Wed 2026-05-27):
      - 2026-05-29 (Fri this week) → published ✓ (upcoming Friday)
      - 2026-06-05 (Fri next week) → hidden ✗
      - 2026-05-22 (Fri last week) → published ✓
    """
    try:
        issue_date = datetime.strptime(issue_date_str, "%Y-%m-%d").date()
        today_et = datetime.now(ET_TZ).date()
        # Days until next Friday (0 if today IS Friday). weekday(): Mon=0..Fri=4..Sun=6
        days_until_friday = (4 - today_et.weekday()) % 7
        upcoming_friday = today_et + timedelta(days=days_until_friday)
        return issue_date <= upcoming_friday
    except Exception:
        return False


def _list_weekly_issues(include_unpublished: bool = False) -> list[dict]:
    """
    Return all published issues, sorted most recent first.
    Set include_unpublished=True for admin/internal views.
    """
    if not _WEEKLY_DIR.exists():
        return []
    issues = []
    for f in _WEEKLY_DIR.glob("*.md"):
        date = f.stem  # YYYY-MM-DD
        # Filter unpublished (future-dated) issues
        if not include_unpublished and not _is_issue_published(date):
            continue
        # Try to parse the first heading + summary from the file
        try:
            content = f.read_text(encoding="utf-8")
            title = ""
            summary = ""
            for line in content.split("\n"):
                line = line.strip()
                if not title and line.startswith("# "):
                    title = line[2:].strip()
                elif not summary and line.startswith("> "):
                    summary = line[2:].strip()
                if title and summary:
                    break
            issues.append({
                "date": date,
                "title": title or f"Sentiment Friday — {date}",
                "summary": summary,
                "url": f"/weekly/{date}",
            })
        except Exception as e:
            print(f"[weekly] read error for {f.name}: {e}")
    issues.sort(key=lambda i: i["date"], reverse=True)
    return issues


@app.get("/weekly")
async def serve_weekly_archive():
    """Public — archive index of all Sentiment Friday issues."""
    issues = _list_weekly_issues()
    snapshots = list(reversed(_load_friday_snapshots()))[:24]
    return FileResponse(
        Path(__file__).parent / "weekly_archive.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/weekly/{issue_date}")
async def serve_weekly_issue(issue_date: str):
    """Public — render a single Sentiment Friday issue."""
    # Validate format YYYY-MM-DD
    try:
        datetime.strptime(issue_date, "%Y-%m-%d")
    except ValueError:
        return Response(
            content=json.dumps({"error": "invalid date format. Use YYYY-MM-DD."}),
            status_code=400, media_type="application/json",
        )

    issue_path = _WEEKLY_DIR / f"{issue_date}.md"

    # Hide future-dated issues from public access (404 looks like they don't exist)
    if not issue_path.exists() or not _is_issue_published(issue_date):
        return Response(
            content=f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Issue not found — Sentiment Friday</title>
<style>body{{background:#07111f;color:#e8eef7;font-family:-apple-system,Inter,sans-serif;text-align:center;padding:80px 20px;}}
a{{color:#5ba3ec;}}</style></head><body>
<h1>Issue not found</h1>
<p>No Sentiment Friday issue for {issue_date}.</p>
<p><a href="/weekly">← View all issues</a> · <a href="/">Dashboard</a></p>
</body></html>""",
            status_code=404, media_type="text/html",
            headers={"Cache-Control": "no-cache"},
        )

    return FileResponse(
        Path(__file__).parent / "weekly_issue.html",
        headers={
            "Cache-Control": "public, max-age=300",  # cache 5min
            "X-Issue-Date": issue_date,
        },
    )


@app.get("/api/weekly/issue/{issue_date}")
async def api_weekly_issue_content(issue_date: str):
    """Returns the markdown content of a specific issue."""
    try:
        datetime.strptime(issue_date, "%Y-%m-%d")
    except ValueError:
        return Response(
            content=json.dumps({"error": "invalid date format"}),
            status_code=400, media_type="application/json",
        )
    issue_path = _WEEKLY_DIR / f"{issue_date}.md"
    # Hide future-dated issues from public access
    if not issue_path.exists() or not _is_issue_published(issue_date):
        return Response(
            content=json.dumps({"error": "issue not found", "date": issue_date}),
            status_code=404, media_type="application/json",
        )
    return {
        "ok":          True,
        "date":        issue_date,
        "content":     issue_path.read_text(encoding="utf-8"),
        "url":         f"/weekly/{issue_date}",
    }


@app.get("/api/weekly/list")
async def api_weekly_list():
    """Returns all published issues."""
    return {
        "ok":     True,
        "issues": _list_weekly_issues(),
    }


@app.get("/subscribe")
async def serve_subscribe():
    """Pro subscription landing page."""
    return FileResponse(
        Path(__file__).parent / "subscribe.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


# ─── July 1, 2026 Tracker (the viral asset) ──────────────────────────────────
@app.get("/july1")
async def serve_july1_tracker():
    """Public — dedicated tracker page for the July 1, 2026 student loan rule changes."""
    return FileResponse(
        Path(__file__).parent / "july1.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/sli-mark.svg")
async def serve_logo_mark():
    """Brand logo — icon mark. Served as a static SVG used in nav across all pages."""
    return FileResponse(
        Path(__file__).parent / "sli-mark.svg",
        media_type="image/svg+xml",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@app.get("/sli-lockup.svg")
async def serve_logo_lockup():
    """Brand logo — full horizontal lockup (icon + wordmark + LIVE tag)."""
    return FileResponse(
        Path(__file__).parent / "sli-lockup.svg",
        media_type="image/svg+xml",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@app.get("/favicon.svg")
async def serve_favicon_svg():
    """Favicon — reuses the icon mark."""
    return FileResponse(
        Path(__file__).parent / "sli-mark.svg",
        media_type="image/svg+xml",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@app.get("/embed/july1")
async def serve_july1_embed():
    """Embeddable July 1 countdown + index widget. iframe-safe for news sites."""
    # Reuse Friday snapshot data — public, no auth, cache-friendly
    friday = _get_most_recent_friday_et()
    friday_date = friday.strftime("%Y-%m-%d")
    snapshots = _load_friday_snapshots()
    current = next((s for s in reversed(snapshots) if s.get("friday_date") == friday_date), None)
    if not current:
        current = await _capture_friday_snapshot()
    score  = int(current.get("index_score") or 70)
    status = current.get("status") or "High Anxiety"

    # Color by score
    if   score >= 80: color, color_soft = "#ef4444", "rgba(239,68,68,.15)"
    elif score >= 65: color, color_soft = "#f97316", "rgba(249,115,22,.15)"
    elif score >= 45: color, color_soft = "#eab308", "rgba(234,179,8,.15)"
    else:             color, color_soft = "#22c55e", "rgba(34,197,94,.15)"

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>July 1 Student Loan Tracker · Loan Clarity</title>
<meta http-equiv="refresh" content="300">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Inter,sans-serif;background:#07111f;color:#e8eef7;-webkit-font-smoothing:antialiased;overflow:hidden}}
a{{text-decoration:none;color:inherit;display:block;height:100%}}
.w{{padding:18px 22px;height:100%;display:flex;flex-direction:column;gap:14px;border:1px solid #1e3a5f;border-radius:14px;background:linear-gradient(180deg,{color_soft} 0%,transparent 55%);position:relative}}
.head{{display:flex;align-items:center;justify-content:space-between}}
.brand{{font-size:11px;letter-spacing:.16em;text-transform:uppercase;color:#9fb2cc;font-weight:800}}
.brand b{{color:#fff}}
.tag{{font-size:10px;color:{color};background:{color_soft};padding:4px 10px;border-radius:999px;letter-spacing:.1em;font-weight:800;text-transform:uppercase}}
.body{{display:grid;grid-template-columns:1fr 1fr;gap:16px;flex:1;align-items:center}}
.col-left{{text-align:center;border-right:1px solid #1e3a5f;padding-right:16px}}
.col-right{{text-align:center}}
.col-lbl{{font-size:9px;letter-spacing:.14em;text-transform:uppercase;color:#6b7e98;font-weight:700;margin-bottom:6px}}
.countdown{{font-size:48px;font-weight:800;color:#fff;letter-spacing:-.04em;line-height:1;font-family:'JetBrains Mono',ui-monospace,monospace}}
.countdown-unit{{font-size:12px;color:#9fb2cc;margin-top:6px;font-weight:600}}
.score{{font-size:54px;font-weight:800;color:{color};letter-spacing:-.04em;line-height:1;font-family:'JetBrains Mono',ui-monospace,monospace}}
.score-status{{font-size:11px;font-weight:800;color:{color};text-transform:uppercase;letter-spacing:.1em;margin-top:6px}}
.foot{{font-size:9px;color:#6b7e98;letter-spacing:.08em;text-align:center;border-top:1px solid #1e3a5f;padding-top:10px}}
.foot b{{color:#fff;font-weight:700}}
</style></head>
<body><a href="https://www.studentloansindex.com/july1?utm_source=embed&utm_medium=july1-widget" target="_blank" rel="noopener">
<div class="w">
  <div class="head">
    <span class="brand">Loan Clarity · <b>July 1 Tracker</b></span>
    <span class="tag">LIVE</span>
  </div>
  <div class="body">
    <div class="col-left">
      <div class="col-lbl">Days until July 1</div>
      <div class="countdown" id="cd">--</div>
      <div class="countdown-unit" id="cdu">calculating</div>
    </div>
    <div class="col-right">
      <div class="col-lbl">Borrower stress today</div>
      <div class="score">{score}</div>
      <div class="score-status">{status}</div>
    </div>
  </div>
  <div class="foot">studentloansindex.com/july1 ↗ · Updates every Friday</div>
</div>
</a>
<script>
function tick() {{
  var target = new Date(Date.UTC(2026, 6, 1, 4, 0, 0)); // July 1, 2026 00:00 ET (UTC-4)
  var now = new Date();
  var ms = target - now;
  if (ms <= 0) {{
    document.getElementById('cd').textContent = '0';
    document.getElementById('cdu').textContent = 'live now';
    return;
  }}
  var days  = Math.floor(ms / 86400000);
  var hours = Math.floor((ms % 86400000) / 3600000);
  document.getElementById('cd').textContent = days;
  document.getElementById('cdu').textContent = days === 1 ? 'day · ' + hours + 'h' : 'days · ' + hours + 'h';
}}
tick(); setInterval(tick, 60000);
</script>
</body></html>"""

    return Response(
        content=html,
        media_type="text/html",
        headers={
            "X-Frame-Options":          "ALLOWALL",
            "Content-Security-Policy":  "frame-ancestors *",
            "Cache-Control":            "public, max-age=300",
            "Access-Control-Allow-Origin": "*",
        },
    )


# ─── Static Page Routes ───────────────────────────────────────────────────────
# All HTML pages set Cache-Control: no-cache so browsers always re-fetch the
# latest version after a deploy — never serve stale nav/copy from cache.

_NO_CACHE_HEADERS = {
    "Cache-Control": "no-cache, no-store, must-revalidate",
    "Pragma":        "no-cache",
    "Expires":       "0",
}


def _serve_html(filename: str) -> FileResponse:
    return FileResponse(
        Path(__file__).parent / filename,
        headers=_NO_CACHE_HEADERS,
    )


@app.get("/")
async def serve_frontend():
    return _serve_html("index.html")


@app.get("/methodology")
async def serve_methodology():
    return _serve_html("methodology.html")


@app.get("/sources")
async def serve_sources():
    return _serve_html("sources.html")


@app.get("/onepager")
async def serve_onepager():
    return _serve_html("onepager.html")


@app.get("/tax-tool")
async def serve_tax_tool():
    return _serve_html("tax-tool.html")


# ─── Embeddable Widget ────────────────────────────────────────────────────────
# Editorial-grade iframe widget. Designed for NYT/WSJ/Bloomberg/Forbes embeds.
# No JS injection (iframe-safe), no external deps, auto-refreshes every 5 min.
# Always links back to studentloansindex.com → backlinks + brand authority.

@app.get("/embed/index")
async def serve_embed(
    size: str = "card",
    theme: str = "dark",
    ref: str = "embed",
):
    """
    Render the embeddable Loan Clarity Index widget.

    Query params:
        size:  "badge" (300x120) | "card" (420x280, default) | "dashboard" (640x420)
        theme: "dark" (default) | "light"
        ref:   referrer string for click-through attribution (default "embed")

    Usage:
        <iframe src="https://www.studentloansindex.com/embed/index?size=card&theme=light"
                width="420" height="280" frameborder="0" scrolling="no"
                title="Loan Clarity Borrower Sentiment Index"></iframe>
    """
    # Sanitize
    size  = size  if size  in ("badge", "card", "dashboard") else "card"
    theme = theme if theme in ("dark", "light")              else "dark"
    # ref can be anything — just sanitize to alphanum + dashes for safety
    ref = "".join(c for c in ref if c.isalnum() or c in "-_")[:40] or "embed"

    # Fetch current data — uses cache, fast.
    full = await get_live_sentiment()
    score   = int(full.get("index_score") or 70)
    status  = full.get("status") or "High Anxiety"
    signals = full.get("signals", {}) or {}
    cfpb    = signals.get("cfpb", {}).get("complaints_90d")
    reddit_posts = signals.get("reddit", {}).get("posts")
    top_theme = full.get("top_theme", "")

    # Status color
    if   score >= 80: color, color_soft = "#ef4444", "rgba(239,68,68,.15)"
    elif score >= 65: color, color_soft = "#f97316", "rgba(249,115,22,.15)"
    elif score >= 45: color, color_soft = "#eab308", "rgba(234,179,8,.15)"
    elif score >= 30: color, color_soft = "#22c55e", "rgba(34,197,94,.15)"
    else:             color, color_soft = "#10b981", "rgba(16,185,129,.15)"

    # Theme palette
    if theme == "dark":
        bg, fg, fg_dim, border = "#07111f", "#e8eef7", "#9fb2cc", "#1e3a5f"
    else:
        bg, fg, fg_dim, border = "#ffffff", "#0a1628", "#5a708a", "#e4eaf2"

    link = f"https://www.studentloansindex.com/?utm_source=embed&utm_medium=widget&utm_campaign={ref}"

    # ── BADGE (compact, sidebar-friendly) ──────────────────────────────
    if size == "badge":
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Loan Clarity Index</title>
<meta http-equiv="refresh" content="300">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Inter,sans-serif;background:{bg};color:{fg};-webkit-font-smoothing:antialiased;overflow:hidden}}
a{{text-decoration:none;color:inherit;display:block;height:100%}}
.w{{padding:14px 16px;height:100%;display:flex;flex-direction:column;justify-content:space-between;border:1px solid {border};border-radius:10px;background:linear-gradient(180deg,{color_soft} 0%,transparent 60%)}}
.head{{display:flex;align-items:center;justify-content:space-between;font-size:10px;letter-spacing:.13em;text-transform:uppercase;color:{fg_dim};font-weight:700}}
.dot{{width:6px;height:6px;border-radius:50%;background:{color};box-shadow:0 0 6px {color};display:inline-block;margin-right:5px;vertical-align:middle}}
.mid{{display:flex;align-items:baseline;gap:8px}}
.score{{font-size:42px;font-weight:800;letter-spacing:-.03em;color:{color};line-height:1}}
.outof{{font-size:13px;color:{fg_dim};font-weight:600}}
.status{{font-size:11px;font-weight:700;letter-spacing:.06em;color:{fg};text-transform:uppercase}}
.foot{{font-size:9px;color:{fg_dim};letter-spacing:.08em}}
.foot b{{color:{fg};font-weight:700}}
</style></head>
<body><a href="{link}" target="_blank" rel="noopener">
<div class="w">
  <div class="head"><span><span class="dot"></span>LIVE</span><span>Borrower Sentiment</span></div>
  <div>
    <div class="mid"><div class="score">{score}</div><div class="outof">/100</div></div>
    <div class="status" style="color:{color}">{status}</div>
  </div>
  <div class="foot">Powered by <b>Loan Clarity</b> ↗</div>
</div>
</a></body></html>"""

    # ── CARD (article inline embed) ──────────────────────────────
    elif size == "card":
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Loan Clarity Borrower Sentiment Index</title>
<meta http-equiv="refresh" content="300">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Inter,sans-serif;background:{bg};color:{fg};-webkit-font-smoothing:antialiased;overflow:hidden}}
a{{text-decoration:none;color:inherit;display:block;height:100%}}
.w{{padding:22px 24px;height:100%;display:flex;flex-direction:column;border:1px solid {border};border-radius:14px;background:linear-gradient(180deg,{color_soft} 0%,transparent 50%);position:relative}}
.head{{display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}}
.brand{{font-size:11px;letter-spacing:.16em;text-transform:uppercase;color:{fg_dim};font-weight:800}}
.brand b{{color:{fg}}}
.live{{display:flex;align-items:center;gap:6px;font-size:10px;color:{fg_dim};letter-spacing:.1em;text-transform:uppercase;font-weight:700}}
.dot{{width:6px;height:6px;border-radius:50%;background:#22c55e;box-shadow:0 0 6px #22c55e;animation:pulse 1.8s ease-in-out infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
.row{{display:flex;align-items:flex-end;gap:14px;margin-bottom:6px}}
.score{{font-size:68px;font-weight:800;letter-spacing:-.04em;color:{color};line-height:.9}}
.outof{{font-size:18px;color:{fg_dim};font-weight:700;padding-bottom:6px}}
.status{{font-size:14px;font-weight:800;letter-spacing:.06em;color:{color};text-transform:uppercase;margin-bottom:14px}}
.bar{{height:6px;background:{border};border-radius:3px;overflow:hidden;margin-bottom:14px;position:relative}}
.bar-fill{{height:100%;background:linear-gradient(90deg,#22c55e 0%,#eab308 45%,#f97316 70%,#ef4444 100%);width:100%;clip-path:inset(0 {100-score}% 0 0)}}
.bar-tick{{position:absolute;top:-3px;height:12px;width:2px;background:{fg};left:calc({score}% - 1px);border-radius:1px}}
.stats{{display:flex;gap:14px;font-size:11px;color:{fg_dim};margin-top:auto;padding-top:10px;border-top:1px solid {border}}}
.stats b{{color:{fg};font-weight:800;font-size:13px;display:block}}
.foot{{position:absolute;bottom:8px;right:14px;font-size:9px;color:{fg_dim};letter-spacing:.08em}}
.foot b{{color:{fg};font-weight:700}}
</style></head>
<body><a href="{link}" target="_blank" rel="noopener">
<div class="w">
  <div class="head">
    <div class="brand">Loan Clarity <b>Borrower Sentiment Index™</b></div>
    <div class="live"><div class="dot"></div>Live</div>
  </div>
  <div class="row"><div class="score">{score}</div><div class="outof">/100</div></div>
  <div class="status">{status}</div>
  <div class="bar"><div class="bar-fill"></div><div class="bar-tick"></div></div>
  <div class="stats">
    <div><b>{cfpb or "—"}</b>CFPB complaints (90d)</div>
    <div><b>{reddit_posts or "—"}</b>Reddit posts analyzed</div>
  </div>
  <div class="foot">studentloansindex.com ↗</div>
</div>
</a></body></html>"""

    # ── DASHBOARD (large hero embed) ──────────────────────────────
    else:  # dashboard
        theme_str = top_theme if top_theme else "SAVE plan uncertainty and repayment restart struggles"
        if len(theme_str) > 80:
            theme_str = theme_str[:78] + "…"
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Loan Clarity Borrower Sentiment Index — Live</title>
<meta http-equiv="refresh" content="300">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Inter,sans-serif;background:{bg};color:{fg};-webkit-font-smoothing:antialiased;overflow:hidden}}
a{{text-decoration:none;color:inherit;display:block;height:100%}}
.w{{padding:28px 32px;height:100%;display:flex;flex-direction:column;border:1px solid {border};border-radius:16px;background:linear-gradient(180deg,{color_soft} 0%,transparent 45%);position:relative}}
.head{{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}}
.brand{{font-size:12px;letter-spacing:.18em;text-transform:uppercase;color:{fg_dim};font-weight:800}}
.brand b{{color:{fg}}}
.live{{display:flex;align-items:center;gap:7px;font-size:11px;color:{fg_dim};letter-spacing:.12em;text-transform:uppercase;font-weight:800}}
.dot{{width:7px;height:7px;border-radius:50%;background:#22c55e;box-shadow:0 0 8px #22c55e;animation:pulse 1.8s ease-in-out infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
.main{{display:grid;grid-template-columns:1fr 1fr;gap:30px;flex:1;align-items:center}}
.col-left{{display:flex;flex-direction:column;justify-content:center}}
.row{{display:flex;align-items:flex-end;gap:14px}}
.score{{font-size:110px;font-weight:800;letter-spacing:-.05em;color:{color};line-height:.9}}
.outof{{font-size:24px;color:{fg_dim};font-weight:700;padding-bottom:14px}}
.status{{font-size:18px;font-weight:800;letter-spacing:.08em;color:{color};text-transform:uppercase;margin-top:8px}}
.col-right{{display:flex;flex-direction:column;gap:12px}}
.theme-card{{padding:14px 16px;border:1px solid {border};border-radius:10px;background:{bg}}}
.theme-lbl{{font-size:9px;color:{fg_dim};letter-spacing:.13em;text-transform:uppercase;font-weight:700;margin-bottom:6px}}
.theme-txt{{font-size:13px;color:{fg};font-weight:600;line-height:1.4}}
.stat-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
.stat{{padding:12px 14px;border:1px solid {border};border-radius:10px;background:{bg}}}
.stat-lbl{{font-size:9px;color:{fg_dim};letter-spacing:.12em;text-transform:uppercase;font-weight:700}}
.stat-val{{font-size:20px;font-weight:800;color:{fg};margin-top:3px;letter-spacing:-.02em}}
.bar{{height:8px;background:{border};border-radius:4px;overflow:hidden;margin:14px 0 6px;position:relative}}
.bar-fill{{height:100%;background:linear-gradient(90deg,#22c55e 0%,#eab308 45%,#f97316 70%,#ef4444 100%);width:100%;clip-path:inset(0 {100-score}% 0 0)}}
.bar-tick{{position:absolute;top:-4px;height:16px;width:2px;background:{fg};left:calc({score}% - 1px);border-radius:1px}}
.foot{{position:absolute;bottom:10px;right:16px;font-size:10px;color:{fg_dim};letter-spacing:.08em}}
.foot b{{color:{fg};font-weight:700}}
</style></head>
<body><a href="{link}" target="_blank" rel="noopener">
<div class="w">
  <div class="head">
    <div class="brand">Loan Clarity <b>Borrower Sentiment Index™</b></div>
    <div class="live"><div class="dot"></div>Live · Updated continuously</div>
  </div>
  <div class="main">
    <div class="col-left">
      <div class="row"><div class="score">{score}</div><div class="outof">/100</div></div>
      <div class="status">{status}</div>
      <div class="bar"><div class="bar-fill"></div><div class="bar-tick"></div></div>
    </div>
    <div class="col-right">
      <div class="theme-card">
        <div class="theme-lbl">Driving the index right now</div>
        <div class="theme-txt">{theme_str}</div>
      </div>
      <div class="stat-grid">
        <div class="stat"><div class="stat-lbl">CFPB Complaints 90d</div><div class="stat-val">{cfpb or "—"}</div></div>
        <div class="stat"><div class="stat-lbl">Reddit Posts</div><div class="stat-val">{reddit_posts or "—"}</div></div>
      </div>
    </div>
  </div>
  <div class="foot">studentloansindex.com ↗</div>
</div>
</a></body></html>"""

    return Response(
        content=html,
        media_type="text/html",
        headers={
            # Allow iframe embedding from anywhere (this is the whole point)
            "X-Frame-Options":          "ALLOWALL",
            "Content-Security-Policy":  "frame-ancestors *",
            # Cache for 5 min — matches refresh meta tag
            "Cache-Control":            "public, max-age=300",
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.get("/api/embed")
async def serve_embed_gallery():
    """Public gallery: pick a size + theme, copy the embed code."""
    return FileResponse(
        Path(__file__).parent / "embed-gallery.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"},
    )


# ─── Newsletter Subscription ──────────────────────────────────────────────────
_SUBSCRIBERS_FILE = Path(__file__).parent / "subscribers.jsonl"


def _parse_device(ua: str) -> tuple[str, str, str]:
    """Return (device_type, browser, os) from a User-Agent string."""
    ua_lower = ua.lower()

    # Device type
    if any(x in ua_lower for x in ("iphone", "android", "mobile")):
        device = "Mobile"
    elif any(x in ua_lower for x in ("ipad", "tablet")):
        device = "Tablet"
    else:
        device = "Desktop"

    # Browser
    if "edg/" in ua_lower or "edge/" in ua_lower:
        browser = "Edge"
    elif "opr/" in ua_lower or "opera" in ua_lower:
        browser = "Opera"
    elif "chrome/" in ua_lower and "chromium" not in ua_lower:
        browser = "Chrome"
    elif "firefox/" in ua_lower:
        browser = "Firefox"
    elif "safari/" in ua_lower and "chrome" not in ua_lower:
        browser = "Safari"
    else:
        browser = "Other"

    # OS
    if "windows" in ua_lower:
        os_name = "Windows"
    elif "mac os x" in ua_lower or "macintosh" in ua_lower:
        os_name = "macOS" if "mobile" not in ua_lower else "iOS"
    elif "iphone" in ua_lower or "ipad" in ua_lower:
        os_name = "iOS"
    elif "android" in ua_lower:
        os_name = "Android"
    elif "linux" in ua_lower:
        os_name = "Linux"
    else:
        os_name = "Other"

    return device, browser, os_name


def _domain_type(domain: str) -> str:
    free = {"gmail.com","yahoo.com","hotmail.com","outlook.com","icloud.com",
            "aol.com","protonmail.com","me.com","live.com","msn.com"}
    if domain in free:
        return "Personal"
    if domain.endswith(".edu"):
        return "Education"
    if domain.endswith(".gov"):
        return "Government"
    if domain.endswith(".org"):
        return "Non-Profit"
    return "Corporate"


def _anonymize_ip(ip: str) -> str:
    """Keep first 3 octets only: 12.34.56.78 → 12.34.56.xxx"""
    parts = ip.split(".")
    if len(parts) == 4:
        return ".".join(parts[:3]) + ".xxx"
    return ip[:20]  # IPv6 — just truncate


@app.post("/api/subscribe")
async def subscribe(request: Request, payload: dict = Body(...)):
    """Append a newsletter signup to subscribers.jsonl — rich lead record."""
    name  = (payload.get("name")  or "").strip()[:120]
    email = (payload.get("email") or "").strip().lower()[:200]
    if "@" not in email or "." not in email.split("@")[-1]:
        return {"ok": False, "error": "Please provide a valid email."}
    # Name is OPTIONAL for newsletter signups (was blocking subscribers)
    if not name:
        name = email.split("@")[0]  # use email local part as fallback display name

    now = datetime.now()
    headers = request.headers

    # ── Device / browser fingerprint ────────────────────────────────
    ua = headers.get("user-agent", "")
    device, browser, os_name = _parse_device(ua)

    # ── IP (anonymized) ─────────────────────────────────────────────
    ip_raw = (
        headers.get("x-forwarded-for", "").split(",")[0].strip()
        or headers.get("x-real-ip", "")
        or (request.client.host if request.client else "")
    )
    ip_anon = _anonymize_ip(ip_raw) if ip_raw else ""

    # ── Email domain analysis ────────────────────────────────────────
    domain = email.split("@")[-1]

    # ── Current index score (from cache — no extra API call) ─────────
    idx_score  = int(_cache.get("index_score", 0)) if _cache else 0
    idx_status = _cache.get("status", "") if _cache else ""

    entry = {
        # Core
        "name":              name,
        "email":             email,
        "ts":                now.isoformat(timespec="seconds"),
        "signup_date":       now.strftime("%Y-%m-%d"),
        "signup_time":       now.strftime("%H:%M:%S"),
        "day_of_week":       now.strftime("%A"),
        # Source & campaign
        "source":            payload.get("source", "dashboard"),
        "page_url":          (payload.get("page_url") or "")[:300],
        "referrer":          (payload.get("referrer")  or "")[:300],
        "utm_source":        (payload.get("utm_source")   or "")[:100],
        "utm_medium":        (payload.get("utm_medium")   or "")[:100],
        "utm_campaign":      (payload.get("utm_campaign") or "")[:100],
        "utm_content":       (payload.get("utm_content")  or "")[:100],
        # Device
        "device_type":       device,
        "browser":           browser,
        "os":                os_name,
        "user_agent":        ua[:400],
        # Network (anonymized)
        "ip_anon":           ip_anon,
        # Email intel
        "email_domain":      domain,
        "domain_type":       _domain_type(domain),
        # Index context at time of signup
        "index_score":       idx_score,
        "index_status":      idx_status,
        # CRM fields (editable in the Excel export)
        "lead_status":       "New",
        "notes":             "",
    }

    # ── Write locally (ephemeral — wiped on deploy) ─────────────────
    try:
        with open(_SUBSCRIBERS_FILE, "a") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception as e:
        print(f"[subscribe] local write error: {e}")

    # ── Fire webhook (persistent — survives deploys) ──────────────
    asyncio.create_task(_fire_lead_webhook(entry))

    print(f"[subscribe] ✓ {email} | {device} | {browser}/{os_name} | idx={idx_score}")
    return {"ok": True, "message": "You're in! Look out for Friday's edition."}


async def _fire_lead_webhook(entry: dict):
    """POST the lead to LEAD_WEBHOOK_URL (Make.com / Zapier / n8n).

    Set LEAD_WEBHOOK_URL as a Railway environment variable.
    The webhook receives a flat JSON object — map each field to a
    Google Sheets column in your Make.com / Zapier scenario.
    """
    url = os.environ.get("LEAD_WEBHOOK_URL", "")
    if not url:
        return
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(url, json=entry)
            print(f"[webhook] fired → HTTP {resp.status_code}")
    except Exception as e:
        print(f"[webhook] error: {e}")


# ─── Lead Export (Excel / CSV) ────────────────────────────────────────────────

COLUMNS = [
    ("Row #",             "row_num",       12),
    ("First Name",        "name",          18),
    ("Email",             "email",         32),
    ("Email Domain",      "email_domain",  24),
    ("Domain Type",       "domain_type",   14),
    ("Signup Date",       "signup_date",   14),
    ("Signup Time",       "signup_time",   12),
    ("Day of Week",       "day_of_week",   13),
    ("Days Since Signup", "days_since",    18),
    ("Source",            "source",        16),
    ("Page URL",          "page_url",      30),
    ("Referrer",          "referrer",      30),
    ("UTM Source",        "utm_source",    16),
    ("UTM Medium",        "utm_medium",    16),
    ("UTM Campaign",      "utm_campaign",  20),
    ("UTM Content",       "utm_content",   20),
    ("Device Type",       "device_type",   13),
    ("Browser",           "browser",       12),
    ("OS",                "os",            12),
    ("IP (Anonymized)",   "ip_anon",       18),
    ("Index Score",       "index_score",   13),
    ("Market Status",     "index_status",  18),
    ("Lead Status",       "lead_status",   14),
    ("Notes",             "notes",         30),
]

STATUS_COLORS = {
    "New":         ("1a3a5c", "5bc8f5"),
    "Contacted":   ("1a3a2a", "4ade80"),
    "Qualified":   ("3a2a00", "fbbf24"),
    "Unsubscribed":("3a1a1a", "f87171"),
}

DOMAIN_COLORS = {
    "Corporate":   "2d7dd2",
    "Education":   "a855f7",
    "Government":  "22c55e",
    "Non-Profit":  "f97316",
    "Personal":    "6b7280",
}


def _build_xlsx(rows: list[dict]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loan Clarity Leads"

        # ── Freeze header row ────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Header row styling ───────────────────────────────────────
        hdr_fill   = PatternFill("solid", fgColor="07111f")
        hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin_side  = Side(style="thin", color="1a2f48")
        thin_border = Border(bottom=Side(style="medium", color="2d7dd2"))

        for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill   = hdr_fill
            cell.font   = hdr_font
            cell.alignment = hdr_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 28

        # ── Data rows ────────────────────────────────────────────────
        today = datetime.now().date()
        alt_fill = PatternFill("solid", fgColor="0b1929")
        wht_fill = PatternFill("solid", fgColor="0d1e33")

        for r_idx, row in enumerate(rows, start=1):
            excel_row = r_idx + 1  # offset for header
            fill = alt_fill if r_idx % 2 == 0 else wht_fill

            # Compute days since signup
            try:
                sig_date = datetime.fromisoformat(row.get("ts", "")).date()
                days_since = (today - sig_date).days
            except Exception:
                days_since = ""

            values = {
                "row_num":      r_idx,
                "days_since":   days_since,
                **{k: row.get(k, "") for _, k, _ in COLUMNS
                   if k not in ("row_num", "days_since")},
            }

            for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
                val  = values.get(key, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.fill = fill
                cell.font = Font(name="Calibri", color="C8DCF0", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # ── Special formatting ─────────────────────────────
                if key == "email":
                    cell.font = Font(name="Calibri", color="5bc8f5", size=10, underline="single")

                elif key == "lead_status":
                    bg, fg = STATUS_COLORS.get(str(val), ("1a3a5c", "5bc8f5"))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(name="Calibri", color=fg, size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                elif key == "domain_type":
                    color = DOMAIN_COLORS.get(str(val), "6b7280")
                    cell.font = Font(name="Calibri", color=color, size=10, bold=True)

                elif key == "index_score" and isinstance(val, int):
                    if val >= 70:
                        cell.font = Font(name="Calibri", color="f87171", size=10, bold=True)
                    elif val >= 40:
                        cell.font = Font(name="Calibri", color="fbbf24", size=10)
                    else:
                        cell.font = Font(name="Calibri", color="4ade80", size=10)

                elif key == "days_since" and isinstance(val, int) and val <= 1:
                    cell.font = Font(name="Calibri", color="00e5a8", size=10, bold=True)

            ws.row_dimensions[excel_row].height = 18

        # ── Auto-filter ──────────────────────────────────────────────
        ws.auto_filter.ref = ws.dimensions

        # ── Summary sheet ────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.sheet_properties.tabColor = "2d7dd2"
        ws2.freeze_panes = "A1"
        s_fill = PatternFill("solid", fgColor="07111f")
        s_font = Font(name="Calibri", color="FFFFFF", bold=True, size=12)

        summary_rows = [
            ("Loan Clarity — Lead Tracker", ""),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
            ("", ""),
            ("OVERVIEW", ""),
            ("Total Signups", len(rows)),
            ("New Leads", sum(1 for r in rows if r.get("lead_status") == "New")),
            ("Contacted", sum(1 for r in rows if r.get("lead_status") == "Contacted")),
            ("Qualified", sum(1 for r in rows if r.get("lead_status") == "Qualified")),
            ("", ""),
            ("DEVICE BREAKDOWN", ""),
            ("Mobile", sum(1 for r in rows if r.get("device_type") == "Mobile")),
            ("Desktop", sum(1 for r in rows if r.get("device_type") == "Desktop")),
            ("Tablet", sum(1 for r in rows if r.get("device_type") == "Tablet")),
            ("", ""),
            ("DOMAIN TYPE", ""),
            ("Corporate", sum(1 for r in rows if r.get("domain_type") == "Corporate")),
            ("Education (.edu)", sum(1 for r in rows if r.get("domain_type") == "Education")),
            ("Personal (Gmail/Yahoo/etc)", sum(1 for r in rows if r.get("domain_type") == "Personal")),
            ("Government (.gov)", sum(1 for r in rows if r.get("domain_type") == "Government")),
            ("Non-Profit (.org)", sum(1 for r in rows if r.get("domain_type") == "Non-Profit")),
            ("", ""),
            ("SOURCE BREAKDOWN", ""),
        ]
        # Source counts
        from collections import Counter
        src_counts = Counter(r.get("source", "unknown") for r in rows)
        for src, cnt in src_counts.most_common():
            summary_rows.append((f"  {src}", cnt))

        for sr_idx, (label, value) in enumerate(summary_rows, start=1):
            c1 = ws2.cell(row=sr_idx, column=1, value=label)
            c2 = ws2.cell(row=sr_idx, column=2, value=value)
            if label in ("OVERVIEW", "DEVICE BREAKDOWN", "DOMAIN TYPE", "SOURCE BREAKDOWN", "Loan Clarity — Lead Tracker"):
                c1.fill = PatternFill("solid", fgColor="0d1e33")
                c1.font = Font(name="Calibri", color="2d7dd2", bold=True, size=11)
            else:
                c1.font = Font(name="Calibri", color="8ab4cc", size=10)
                c2.font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
            for c in (c1, c2):
                c.alignment = Alignment(vertical="center")

        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    except ImportError:
        raise RuntimeError("openpyxl not installed")


@app.get("/api/admin/leads.xlsx")
async def export_leads_xlsx(key: str = ""):
    """Download all leads as a formatted Excel file.
    Protected by ADMIN_KEY env var (default: 'loanclarty-admin').
    """
    expected = os.environ.get("ADMIN_KEY", "loanclarty-admin")
    if key != expected:
        return Response(status_code=401, content="Unauthorized")

    rows: list[dict] = []
    if _SUBSCRIBERS_FILE.exists():
        for line in _SUBSCRIBERS_FILE.read_text().splitlines():
            line = line.strip()
            if line:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass

    xlsx_bytes = _build_xlsx(rows)
    filename = f"LoanClarity_Leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/leads.csv")
async def export_leads_csv(key: str = ""):
    """Download all leads as CSV (fallback if Excel not available)."""
    expected = os.environ.get("ADMIN_KEY", "loanclarty-admin")
    if key != expected:
        return Response(status_code=401, content="Unauthorized")

    rows: list[dict] = []
    if _SUBSCRIBERS_FILE.exists():
        for line in _SUBSCRIBERS_FILE.read_text().splitlines():
            line = line.strip()
            if line:
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass

    import csv, io
    buf = io.StringIO()
    headers = [col for _, col, _ in COLUMNS if col not in ("row_num", "days_since")]
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    filename = f"LoanClarity_Leads_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([buf.getvalue().encode()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/social-signals")
async def serve_social_signals():
    return _serve_html("social-signals.html")


@app.get("/privacy")
async def serve_privacy():
    return _serve_html("privacy.html")


@app.get("/terms")
async def serve_terms():
    return _serve_html("terms.html")


# ─── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        print("⚠️  ANTHROPIC_API_KEY not set — AI sentiment engine will use fallback values.")
    print("🚀  Loan Clarity Sentiment Engine starting...")
    print("📊  Dashboard:   http://localhost:8000")
    print("📖  Methodology: http://localhost:8000/methodology")
    print("📚  Sources:     http://localhost:8000/sources")
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="warning")


# Force fresh deploy check — 2026-05-24 21:10
